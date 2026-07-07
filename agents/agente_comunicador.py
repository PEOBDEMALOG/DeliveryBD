# peo_bd/agents/agente_comunicador.py
# Agente 4 — Comunicador
# Responsabilidade: para cada onda do plano, gera a planilha de programação
# de coleta e envia por e-mail à transportadora. Registra confirmação e protocolo.

import io
import logging
import smtplib
import uuid
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from sqlalchemy.orm import selectinload

from core.config import settings
from core.historico import HistoricoService
from core.models import Onda, PlanoDia, ProgramacaoColeta, Remessa

logger = logging.getLogger(__name__)


class AgenteComunicador:
    """
    Agente 4 — Comunicador.

    Fluxo:
      1. Recebe plano_id (após Agente 3 concluir)
      2. Para cada onda do plano:
         a. Gera planilha Excel de programação (formato que a DHL/UPS já reconhece)
         b. Monta e-mail estruturado
         c. Envia via SMTP (ou simula em modo dry_run)
         d. Registra ProgramacaoColeta no banco
      3. Retorna resumo de envios
    """

    def __init__(self, db: AsyncSession, dry_run: bool = False):
        self.db       = db
        self.dry_run  = dry_run  # True = não envia e-mail, só registra
        self.historico = HistoricoService(db)

    # ── Entrada principal ──────────────────────────────────────────────────────

    async def programar_coletas(self, plano_id: int) -> dict[str, Any]:
        logger.info(f"[Comunicador] Iniciando programação para plano_id={plano_id}")

        plano = await self._get_plano(plano_id)
        if not plano:
            raise ValueError(f"Plano não encontrado: {plano_id}")

        ondas = await self._carregar_ondas(plano_id)
        enviados = erros = 0

        for onda in ondas:
            try:
                await self._programar_onda(onda, plano)
                enviados += 1
            except Exception as e:
                erros += 1
                logger.error(f"[Comunicador] Erro na onda {onda.id}: {e}")

        await self.db.commit()

        resultado = {
            "plano_id": plano_id,
            "ondas":    len(ondas),
            "enviados": enviados,
            "erros":    erros,
            "dry_run":  self.dry_run,
        }
        logger.info(f"[Comunicador] {enviados} programações enviadas, {erros} erros")
        return resultado

    # ── Programação por onda ───────────────────────────────────────────────────

    async def _programar_onda(self, onda: Onda, plano: PlanoDia) -> None:
        if not onda.transportadora:
            logger.warning(f"[Comunicador] Onda {onda.id} sem transportadora — pulando")
            return

        remessas = await self._carregar_remessas_onda(onda.id)
        if not remessas:
            return

        # Gera planilha Excel
        xlsx_bytes, xlsx_nome = self._gerar_xlsx(onda, remessas, plano)

        # Salva anexo em disco
        output_dir = settings.OUTPUT_DIR / str(plano.data_plano)
        output_dir.mkdir(parents=True, exist_ok=True)
        xlsx_path = output_dir / xlsx_nome
        xlsx_path.write_bytes(xlsx_bytes)

        # Monta corpo do e-mail
        assunto = self._assunto_email(onda, plano)
        corpo   = self._corpo_email(onda, remessas, plano)
        dest    = onda.transportadora.email_operacoes or ""

        # Cria registro de programação
        prog = ProgramacaoColeta(
            onda_id             = onda.id,
            transportadora_id   = onda.transportadora_id,
            canal               = "email",
            destinatario_email  = dest,
            assunto             = assunto,
            corpo               = corpo,
            arquivo_anexo       = str(xlsx_path),
            status_envio        = "pendente",
        )
        self.db.add(prog)
        await self.db.flush()

        # Envia (ou simula)
        if self.dry_run:
            logger.info(f"[Comunicador] dry_run — e-mail simulado para {dest}: {assunto}")
            prog.status_envio = "simulado"
            prog.protocolo    = f"DRY-{uuid.uuid4().hex[:8].upper()}"
            prog.enviado_em   = datetime.utcnow()
            await self.historico.registrar(
                tipo_evento="decisao_agente",
                origem="comunicador",
                ator_tipo="agente_ia",
                ator_nome="Agente Comunicador",
                transportadora_id=onda.transportadora_id,
                descricao=f"Programação simulada (dry_run) — {onda.nome} → {onda.transportadora.nome} <{dest}>",
                resultado="acao_automatica",
                gravidade="info",
                dados_extra={"onda_id": onda.id, "protocolo": prog.protocolo, "dry_run": True},
            )
        else:
            try:
                self._enviar_email(dest, assunto, corpo, xlsx_bytes, xlsx_nome)
                prog.status_envio = "enviado"
                prog.protocolo    = f"BD-{uuid.uuid4().hex[:8].upper()}"
                prog.enviado_em   = datetime.utcnow()
                logger.info(f"[Comunicador] E-mail enviado → {dest} | protocolo: {prog.protocolo}")
                await self.historico.registrar(
                    tipo_evento="decisao_agente",
                    origem="comunicador",
                    ator_tipo="agente_ia",
                    ator_nome="Agente Comunicador",
                    transportadora_id=onda.transportadora_id,
                    descricao=f"Programação enviada — {onda.nome} → {onda.transportadora.nome} <{dest}> | protocolo {prog.protocolo}",
                    resultado="sucesso",
                    dados_extra={"onda_id": onda.id, "protocolo": prog.protocolo, "destinatario": dest},
                )
            except Exception as e:
                prog.status_envio = "erro"
                logger.error(f"[Comunicador] Falha ao enviar e-mail: {e}")
                await self.historico.registrar(
                    tipo_evento="erro_sistema",
                    origem="comunicador",
                    ator_tipo="agente_ia",
                    ator_nome="Agente Comunicador",
                    transportadora_id=onda.transportadora_id,
                    descricao=f"Falha ao enviar e-mail para {onda.transportadora.nome} ({dest}): {e}",
                    resultado="falha",
                    gravidade="alerta",
                    dados_extra={"onda_id": onda.id, "erro": str(e)},
                )
                raise

    # ── Geração de Excel ──────────────────────────────────────────────────────

    def _gerar_xlsx(
        self,
        onda: Onda,
        remessas: list[Remessa],
        plano: PlanoDia,
    ) -> tuple[bytes, str]:
        wb = Workbook()
        ws = wb.active
        ws.title = f"Onda {onda.numero_onda:02d}"

        # Estilos
        header_fill = PatternFill("solid", fgColor="1A3F6F")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        info_font   = Font(bold=True, size=10)

        # Bloco de cabeçalho
        ws.merge_cells("A1:H1")
        ws["A1"] = f"PROGRAMAÇÃO DE COLETA — BECTON DICKINSON — {plano.data_plano.strftime('%d/%m/%Y')}"
        ws["A1"].font = Font(bold=True, size=13, color="1A3F6F")

        ws["A2"] = "Transportadora:"
        ws["B2"] = onda.transportadora.nome if onda.transportadora else ""
        ws["C2"] = "Onda:"
        ws["D2"] = onda.nome
        ws["E2"] = "Tipo:"
        ws["F2"] = (onda.tipo or "").upper()
        ws["G2"] = "Horário coleta:"
        ws["H2"] = onda.horario_coleta.strftime("%Hh%M") if onda.horario_coleta else "A definir"

        ws["A3"] = "Volume total (m³):"
        ws["B3"] = float(onda.volume_total_m3 or 0)
        ws["C3"] = "Peso total (kg):"
        ws["D3"] = float(onda.peso_total_kg or 0)
        ws["E3"] = "Valor NF total (R$):"
        ws["F3"] = float(onda.valor_total_nf or 0)
        ws["G3"] = "Ocupação veículo:"
        ws["H3"] = f"{float(onda.ocupacao_pct or 0):.1f}%"

        for cell in ["A2", "C2", "E2", "G2", "A3", "C3", "E3", "G3"]:
            ws[cell].font = info_font

        # Linha em branco
        ws.append([])

        # Cabeçalho da tabela
        colunas = [
            "Nº Remessa", "Nº NF", "Destinatário", "Cidade/UF",
            "Volume (m³)", "Peso (kg)", "Valor NF (R$)",
            "Janela Entrega", "Prioridade", "Tipo"
        ]
        ws.append(colunas)
        for col_num, _ in enumerate(colunas, start=1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Dados
        for seq, r in enumerate(remessas, start=1):
            cliente = r.cliente
            cidade_uf = f"{cliente.cidade or ''}/{cliente.uf or ''}" if cliente else ""
            janela = ""
            if r.janela_inicio and r.janela_fim:
                janela = f"{r.janela_inicio.strftime('%Hh%M')} – {r.janela_fim.strftime('%Hh%M')}"

            row_data = [
                r.numero_remessa,
                r.numero_nf or "Pendente",
                cliente.razao_social if cliente else "",
                cidade_uf,
                float(r.volume_m3 or 0),
                float(r.peso_kg or 0),
                float(r.valor_nf or 0),
                janela or "Flexível",
                (r.prioridade or "normal").upper(),
                (r.tipo_entrega or onda.tipo or "").upper(),
            ]
            ws.append(row_data)

            # Destaca ATA em azul claro
            if r.is_ata:
                ata_fill = PatternFill("solid", fgColor="D5E8F4")
                for col in range(1, len(row_data) + 1):
                    ws.cell(row=ws.max_row, column=col).fill = ata_fill

            # Destaca NF pendente em amarelo
            if not r.nf_emitida:
                nf_fill = PatternFill("solid", fgColor="FAEEDA")
                ws.cell(row=ws.max_row, column=2).fill = nf_fill

        # Largura das colunas
        larguras = [14, 14, 35, 18, 12, 12, 16, 18, 12, 12]
        for i, larg in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(i)].width = larg

        # Salva em memória
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        nome = (
            f"Prog_{onda.transportadora.codigo if onda.transportadora else 'TRANSP'}"
            f"_Onda{onda.numero_onda:02d}"
            f"_{plano.data_plano.strftime('%d%m%Y')}.xlsx"
        )
        return buf.read(), nome

    # ── E-mail ────────────────────────────────────────────────────────────────

    def _assunto_email(self, onda: Onda, plano: PlanoDia) -> str:
        transp = onda.transportadora.codigo if onda.transportadora else "TRANSP"
        cd     = onda.plano.cd.codigo if onda.plano and onda.plano.cd else "CD"
        return (
            f"Programação BD {cd} — {onda.nome} — "
            f"{plano.data_plano.strftime('%d/%m/%Y')}"
        )

    def _corpo_email(
        self, onda: Onda, remessas: list[Remessa], plano: PlanoDia
    ) -> str:
        transp_nome = onda.transportadora.nome if onda.transportadora else "Transportadora"
        cd_nome     = onda.plano.cd.nome if onda.plano and onda.plano.cd else "CD"
        horario     = onda.horario_coleta.strftime("%Hh%M") if onda.horario_coleta else "a confirmar"

        nf_pendentes = [r for r in remessas if not r.nf_emitida]
        atas_criticas = [r for r in remessas if r.is_ata and r.dias_restantes and r.dias_restantes <= 3]

        linhas_atas = ""
        if atas_criticas:
            linhas_atas = "\n⚠ ATENÇÃO — EMPENHOS ATA COM PRAZO CRÍTICO:\n"
            for r in atas_criticas:
                linhas_atas += (
                    f"  • {r.numero_remessa} → {r.cliente.razao_social if r.cliente else 'N/D'} "
                    f"— vence em {r.prazo_empenho.strftime('%d/%m')} ({r.dias_restantes} dia(s))\n"
                )

        linhas_nf = ""
        if nf_pendentes:
            linhas_nf = "\n⚠ REMESSAS COM NF PENDENTE (aguardar antes de gerar etiqueta):\n"
            for r in nf_pendentes:
                linhas_nf += f"  • {r.numero_remessa} — {r.cliente.razao_social if r.cliente else 'N/D'}\n"

        corpo = f"""Prezados {transp_nome},

Segue programação de coleta Becton Dickinson — {cd_nome}:

DATA: {plano.data_plano.strftime('%d/%m/%Y')}
ONDA: {onda.nome}
TIPO: {(onda.tipo or '').upper()}
HORÁRIO DE COLETA SUGERIDO: {horario}

RESUMO DO LOTE:
  • Remessas: {len(remessas)}
  • Volume total: {float(onda.volume_total_m3 or 0):.2f} m³
  • Peso total: {float(onda.peso_total_kg or 0):.1f} kg
  • Valor NF total: R$ {float(onda.valor_total_nf or 0):,.2f}
  • Ocupação veículo: {float(onda.ocupacao_pct or 0):.1f}%
{linhas_atas}{linhas_nf}
Detalhe completo no arquivo anexo.

Favor confirmar veículo designado e horário real de coleta respondendo este e-mail com o protocolo gerado.

Atenciosamente,
Sistema PEO-BD — Becton Dickinson
(gerado automaticamente em {datetime.now().strftime('%d/%m/%Y %Hh%M')})"""
        return corpo

    def _enviar_email(
        self,
        destinatario: str,
        assunto: str,
        corpo: str,
        xlsx_bytes: bytes,
        xlsx_nome: str,
    ) -> None:
        msg = MIMEMultipart()
        msg["From"]    = settings.EMAIL_FROM
        msg["To"]      = destinatario
        msg["Subject"] = assunto
        msg.attach(MIMEText(corpo, "plain", "utf-8"))

        anexo = MIMEApplication(xlsx_bytes, _subtype="xlsx")
        anexo.add_header("Content-Disposition", "attachment", filename=xlsx_nome)
        msg.attach(anexo)

        with smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT) as server:
            server.ehlo()
            server.starttls()
            server.login(settings.SMTP_USER, settings.SMTP_PASSWORD)
            server.send_message(msg)

    # ── Queries ───────────────────────────────────────────────────────────────

    async def _get_plano(self, plano_id: int) -> PlanoDia | None:
        res = await self.db.execute(
            select(PlanoDia)
            .options(selectinload(PlanoDia.cd))
            .where(PlanoDia.id == plano_id)
        )
        return res.scalar_one_or_none()

    async def _carregar_ondas(self, plano_id: int) -> list[Onda]:
        res = await self.db.execute(
            select(Onda)
            .options(
                selectinload(Onda.transportadora),
                selectinload(Onda.veiculo),
                selectinload(Onda.plano).selectinload(PlanoDia.cd),
            )
            .where(Onda.plano_id == plano_id)
        )
        return list(res.scalars().all())

    async def _carregar_remessas_onda(self, onda_id: int) -> list[Remessa]:
        from core.models import OndaRemessa
        res = await self.db.execute(
            select(Remessa)
            .options(selectinload(Remessa.cliente))
            .join(OndaRemessa, OndaRemessa.remessa_id == Remessa.id)
            .where(OndaRemessa.onda_id == onda_id)
            .order_by(OndaRemessa.sequencia)
        )
        return list(res.scalars().all())
