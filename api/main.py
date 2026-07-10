# peo_bd/api/main.py
# API REST — FastAPI
# Expõe todos os agentes via HTTP para o painel web e integrações futuras.

import hashlib
import json
import logging
import shutil
from datetime import date, datetime, timedelta, time as dtime
from pathlib import Path
from typing import Any, List, Optional

from fastapi import (
    Depends, FastAPI, File, Form, HTTPException, Request,
    UploadFile, WebSocket, WebSocketDisconnect
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func, update, exists
from sqlalchemy.orm import selectinload, joinedload

from core.config import settings, IS_VERCEL
from core.cache import get_cached, set_cached
from core.models import (
    Base, Alerta, Remessa, PlanoDia, Onda, OndaRemessa,
    OportunidadeConsolidacao, CentroDistribuicao,
    Cliente, Transportadora, TabelaPrecoTransportadora, Upload,
    HistoricoEventos, ProgramacaoColeta, TipoErro, ErroAcao,
)
from agents.orquestrador import Orquestrador
from agents.agente_resolvedor import AgenteResolvedor
from agents.agente_classificador import AgenteClassificador
from agents.agente_montador import AgenteMontador

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(name)s | %(levelname)s | %(message)s",
)


def serializar_evento(e: HistoricoEventos) -> dict:
    return {
        "id":                e.id,
        "timestamp":         e.timestamp.isoformat() if e.timestamp else None,
        "tipo_evento":       e.tipo_evento,
        "origem":            e.origem,
        "ator_tipo":         e.ator_tipo,
        "ator_nome":         e.ator_nome,
        "remessa_id":        e.remessa_id,
        "transportadora_id": e.transportadora_id,
        "cd_id":             e.cd_id,
        "descricao":         e.descricao,
        "resultado":         e.resultado,
        "gravidade":         e.gravidade,
        "visibilidade":      e.visibilidade,
        "dados_extra":       e.dados_extra,
    }


logger = logging.getLogger(__name__)

BASE_DIR     = Path(__file__).resolve().parent.parent
FRONTEND_DIR = BASE_DIR / "frontend"

# ── App ───────────────────────────────────────────────────────────────────────

app = FastAPI(
    title       = "PEO-BD — Planejamento de Expedição Outbound",
    description = "Agente robô logístico — Becton Dickinson",
    version     = "1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins     = ["*"],
    allow_credentials = True,
    allow_methods     = ["*"],
    allow_headers     = ["*"],
)

# Rotas de /api/* isentas de JWT: login (ainda não há token), health check do
# Vercel/monitoramento, e o Cron do Resolvedor (já protegido por CRON_SECRET).
ROTAS_PUBLICAS = {"/api/auth/login", "/api/ping", "/api/resolvedor/executar"}


@app.middleware("http")
async def exigir_jwt(request: Request, call_next):
    from jose import JWTError
    from core.auth import verificar_token

    path = request.url.path
    if path.startswith("/api/") and path not in ROTAS_PUBLICAS:
        auth = request.headers.get("authorization", "")
        token = auth.replace("Bearer ", "") if auth.startswith("Bearer ") else ""
        try:
            verificar_token(token)
        except JWTError:
            return JSONResponse(status_code=401, content={"detail": "Não autenticado"})

    return await call_next(request)

# ── Banco ─────────────────────────────────────────────────────────────────────
# Engine/sessionmaker/get_db vivem em core/db.py (compartilhado com agents/ para
# permitir sessões concorrentes extras em leituras paralelizáveis).
from core.db import engine, AsyncSessionLocal, get_db  # noqa: E402


async def get_orq(db: AsyncSession = Depends(get_db)) -> Orquestrador:
    return Orquestrador(db)


# ── AUTENTICAÇÃO ────────────────────────────────────────────────────────────────

class LoginSchema(BaseModel):
    usuario: str
    senha:   str


@app.post("/api/auth/login", summary="Login — retorna token JWT (usuários fictícios de demo)")
async def login(body: LoginSchema):
    from core.auth import USUARIOS, criar_token

    user = USUARIOS.get(body.usuario)
    if not user or user["senha"] != body.senha:
        raise HTTPException(status_code=401, detail="Credenciais inválidas")

    token = criar_token(body.usuario)
    return {"token": token, "nome": user["nome"], "cd": user["cd"], "role": user["role"]}


@app.get("/api/auth/me", summary="Retorna os dados do usuário autenticado a partir do token")
async def me(request: Request):
    from jose import JWTError
    from core.auth import verificar_token

    auth  = request.headers.get("authorization", "")
    token = auth.replace("Bearer ", "") if auth.startswith("Bearer ") else ""
    try:
        return verificar_token(token)
    except JWTError:
        raise HTTPException(status_code=401, detail="Token inválido ou expirado")


@app.on_event("startup")
async def startup():
    for _dir in [settings.UPLOAD_DIR, settings.OUTPUT_DIR]:
        try:
            _dir.mkdir(parents=True, exist_ok=True)
        except OSError:
            pass
    if "sqlite" in settings.DATABASE_URL:
        try:
            settings.DB_DIR.mkdir(parents=True, exist_ok=True)
        except OSError:
            pass
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
        await _garantir_colunas_novas(conn)
    await _seed_dados_base()
    logger.info(f"PEO-BD iniciado | db={settings.DATABASE_URL} | vercel={IS_VERCEL} | db_dir={settings.DB_DIR}")


async def _garantir_colunas_novas(conn):
    """create_all() só cria tabelas que faltam — não altera tabelas já existentes.
    Sem migrations formais no projeto, colunas adicionadas após o primeiro deploy
    entram aqui via ALTER TABLE idempotente (ignora erro se a coluna já existe,
    já que SQLite não suporta ADD COLUMN IF NOT EXISTS)."""
    from sqlalchemy import text
    is_sqlite = "sqlite" in settings.DATABASE_URL
    alteracoes = [
        ("ondas", "justificativa", "TEXT"),
        ("transportadoras", "meta_otif", "FLOAT DEFAULT 95.0"),
        ("remessas", "motivo_tentativa", "VARCHAR(200)"),
    ]
    for tabela, coluna, tipo in alteracoes:
        stmt = (
            f"ALTER TABLE {tabela} ADD COLUMN {coluna} {tipo}"
            if is_sqlite else
            f"ALTER TABLE {tabela} ADD COLUMN IF NOT EXISTS {coluna} {tipo}"
        )
        try:
            await conn.execute(text(stmt))
        except Exception:
            pass  # coluna já existe


async def _seed_dados_base():
    """Garante infraestrutura (CDs, transportadoras, veículos, clientes).
    Executado no startup — idempotente (só insere se estiver vazio).
    Remessas entram pelos arquivos de upload — o dashboard começa zerado."""
    from core.models import Veiculo, TabelaPrecoTransportadora

    async with AsyncSessionLocal() as db:
        res = await db.execute(select(CentroDistribuicao))
        if res.scalars().first() is not None:
            return  # já seedado

        logger.info("[Startup] Banco vazio — criando infraestrutura base...")

        # ── CDs ───────────────────────────────────────────────────────────────
        cd_osa = CentroDistribuicao(codigo="OSA", nome="CD Osasco", cidade="Osasco",
                                    uf="SP", sistema_origem="SAP",     capacidade_dia=10000)
        cd_itj = CentroDistribuicao(codigo="ITJ", nome="CD Itajaí",  cidade="Itajaí",
                                    uf="SC", sistema_origem="UPS_WMS", capacidade_dia=5000)
        db.add_all([cd_osa, cd_itj])
        await db.flush()

        # ── Transportadoras ───────────────────────────────────────────────────
        # meta_otif: contrato premium/frota própria toleram menos desvio que fracionado.
        t_dhl    = Transportadora(codigo="DHL",      nome="DHL Express Brasil",    cd_id=cd_osa.id,
                                  email_operacoes="operacoes.sp@dhl.com",    integracao="email",   sla_resposta_h=2, meta_otif=97.0)
        t_ups    = Transportadora(codigo="UPS",      nome="UPS Brasil",            cd_id=cd_itj.id,
                                  email_operacoes="coletas.sc@ups.com",      integracao="email",   sla_resposta_h=2, meta_otif=96.0)
        t_frota  = Transportadora(codigo="FROTA_BD", nome="Frota Própria BD SP",   cd_id=cd_osa.id,
                                  email_operacoes=None,                      integracao="interno", sla_resposta_h=0, meta_otif=98.0)
        t_jadlog = Transportadora(codigo="JADLOG",   nome="Jadlog Logística",      cd_id=cd_osa.id,
                                  email_operacoes="coletas.sp@jadlog.com.br",integracao="email",   sla_resposta_h=4, meta_otif=94.0)
        t_tnt    = Transportadora(codigo="TNT",      nome="TNT Mercúrio (FedEx)",  cd_id=cd_osa.id,
                                  email_operacoes="operacoes@tnt.com.br",   integracao="email",   sla_resposta_h=3, meta_otif=95.0)
        db.add_all([t_dhl, t_ups, t_frota, t_jadlog, t_tnt])
        await db.flush()

        # ── Veículos ──────────────────────────────────────────────────────────
        veiculos = [
            Veiculo(cd_id=cd_osa.id, tipo="truck",         placa="ABC-1234", proprietario="dhl",           capacidade_m3=30.0, capacidade_kg=10000.0),
            Veiculo(cd_id=cd_osa.id, tipo="vuc_eletrico",  placa="ELT-0001", proprietario="frota_propria", capacidade_m3=20.0, capacidade_kg=3500.0),
            Veiculo(cd_id=cd_osa.id, tipo="vuc_combustao", placa="VUC-5678", proprietario="frota_propria", capacidade_m3=10.0, capacidade_kg=2000.0),
            Veiculo(cd_id=cd_osa.id, tipo="van",           placa="VAN-0001", proprietario="frota_propria", capacidade_m3=3.0,  capacidade_kg=600.0),
            Veiculo(cd_id=cd_osa.id, tipo="van",           placa="VAN-0002", proprietario="frota_propria", capacidade_m3=3.0,  capacidade_kg=600.0),
            Veiculo(cd_id=cd_itj.id, tipo="truck",         placa="UPS-9900", proprietario="ups",           capacidade_m3=30.0, capacidade_kg=10000.0),
        ]
        db.add_all(veiculos)
        await db.flush()

        # ── Tabelas de preço ──────────────────────────────────────────────────
        precos = []
        for t_id, regioes in [
            (t_dhl.id,    [("Sudeste","São Paulo","SP","Capital",   12.0,180.0,4200.0,1),
                           ("Sudeste","São Paulo","SP","Interior",   9.0,250.0,6500.0,2),
                           ("Sul","Santa Catarina","SC","Interior", 15.0,350.0,8800.0,3)]),
            (t_jadlog.id, [("Sudeste","São Paulo","SP","Capital",    8.0,120.0,3200.0,2),
                           ("Sudeste","São Paulo","SP","Interior",   7.0,200.0,5200.0,3),
                           ("Sul","Santa Catarina","SC","Interior", 11.0,280.0,7200.0,4)]),
            (t_tnt.id,    [("Sudeste","São Paulo","SP","Capital",   10.0,150.0,3800.0,1),
                           ("Sudeste","São Paulo","SP","Interior",   8.0,220.0,5800.0,2),
                           ("Sul","Santa Catarina","SC","Interior", 13.0,300.0,8000.0,3)]),
        ]:
            for macro, estado, uf, classif, pkg, pmin, pftl, prazo in regioes:
                precos.append(TabelaPrecoTransportadora(
                    transportadora_id=t_id, macro_regiao=macro, estado=estado,
                    uf=uf, classificacao=classif, cobertura=True,
                    preco_por_kg=pkg, preco_minimo=pmin, prazo_frac_dias=prazo,
                    preco_ftl_fixo=pftl, prazo_ftl_dias=prazo,
                    ad_valorem_pct=0.0015, gris_pct=0.001, ativo=True,
                ))
        db.add_all(precos)
        await db.flush()

        # ── Clientes base ─────────────────────────────────────────────────────
        from core.models import Cliente
        clientes_data = [
            dict(razao_social="Hospital das Clínicas FMUSP",     tipo="hospital",    cidade="São Paulo",     uf="SP", regiao="capital_sp",  tem_armazenagem=False, janela_flexivel=False, contrato_ata=True,  prazo_ata_dias=30, volume_medio_m3=1.5),
            dict(razao_social="Hospital Oswaldo Cruz",            tipo="hospital",    cidade="São Paulo",     uf="SP", regiao="capital_sp",  tem_armazenagem=False, janela_flexivel=False, contrato_ata=True,  prazo_ata_dias=30, volume_medio_m3=1.2),
            dict(razao_social="Hospital Albert Einstein",         tipo="hospital",    cidade="São Paulo",     uf="SP", regiao="capital_sp",  tem_armazenagem=True,  janela_flexivel=False, contrato_ata=True,  prazo_ata_dias=45, volume_medio_m3=3.0),
            dict(razao_social="Hospital Sírio-Libanês",           tipo="hospital",    cidade="São Paulo",     uf="SP", regiao="capital_sp",  tem_armazenagem=True,  janela_flexivel=False, contrato_ata=True,  prazo_ata_dias=45, volume_medio_m3=2.8),
            dict(razao_social="Fleury Medicina Diagnóstica SP",   tipo="laboratorio", cidade="São Paulo",     uf="SP", regiao="capital_sp",  tem_armazenagem=True,  janela_flexivel=True,  contrato_ata=False, prazo_ata_dias=None, volume_medio_m3=2.5),
            dict(razao_social="Hermes Pardini SP",                tipo="laboratorio", cidade="São Paulo",     uf="SP", regiao="capital_sp",  tem_armazenagem=True,  janela_flexivel=True,  contrato_ata=False, prazo_ata_dias=None, volume_medio_m3=3.5),
            dict(razao_social="Hospital Municipal de Campinas",   tipo="hospital",    cidade="Campinas",      uf="SP", regiao="interior_sp", tem_armazenagem=False, janela_flexivel=False, contrato_ata=True,  prazo_ata_dias=30, volume_medio_m3=1.0),
            dict(razao_social="Hospital das Clínicas UNICAMP",    tipo="hospital",    cidade="Campinas",      uf="SP", regiao="interior_sp", tem_armazenagem=True,  janela_flexivel=False, contrato_ata=True,  prazo_ata_dias=30, volume_medio_m3=2.5),
            dict(razao_social="Hospital Regional de Itajaí",      tipo="hospital",    cidade="Itajaí",        uf="SC", regiao="sul",         tem_armazenagem=False, janela_flexivel=False, contrato_ata=True,  prazo_ata_dias=30, volume_medio_m3=1.5),
            dict(razao_social="Hospital Municipal de Blumenau",   tipo="hospital",    cidade="Blumenau",      uf="SC", regiao="sul",         tem_armazenagem=False, janela_flexivel=False, contrato_ata=True,  prazo_ata_dias=30, volume_medio_m3=1.2),
            dict(razao_social="UFSC Hospital Universitário",      tipo="universidade",cidade="Florianópolis", uf="SC", regiao="sul",         tem_armazenagem=True,  janela_flexivel=True,  contrato_ata=True,  prazo_ata_dias=60, volume_medio_m3=3.5),
        ]
        clientes_obj = []
        for c in clientes_data:
            obj = Cliente(
                razao_social=c["razao_social"], tipo=c["tipo"],
                cidade=c["cidade"], uf=c["uf"], regiao=c["regiao"],
                tem_armazenagem=c["tem_armazenagem"], janela_flexivel=c["janela_flexivel"],
                contrato_ata=c["contrato_ata"], prazo_ata_dias=c.get("prazo_ata_dias"),
                volume_medio_m3=c["volume_medio_m3"], perfil_volume="fracionado",
            )
            db.add(obj)
            clientes_obj.append(obj)
        await db.flush()

        await db.commit()
        logger.info("[Startup] Seed concluído: 2 CDs, 5 transportadoras, 6 veículos, 11 clientes base")


# ── UPLOAD E PIPELINE COMPLETO ────────────────────────────────────────────────

@app.post("/api/upload", summary="Upload de arquivo + pipeline completo")
async def upload_arquivo(
    arquivo:     UploadFile = File(...),
    cd_codigo:   str        = Form(..., description="'OSA' ou 'ITJ'"),
    usuario:     str        = Form(..., description="'timoteo' ou 'carlos'"),
    origem:      str        = Form(None, description="'SAP' ou 'UPS_WMS' (opcional)"),
    auto_enviar: bool       = Form(True),
    orq:         Orquestrador = Depends(get_orq),
):
    destino = settings.UPLOAD_DIR / f"{cd_codigo}_{usuario}_{arquivo.filename}"
    destino.write_bytes(await arquivo.read())

    logger.info(f"[API] Upload recebido: {arquivo.filename} | CD={cd_codigo} | user={usuario}")

    try:
        resultado = await orq.processar_arquivo(
            arquivo_path = destino,
            cd_codigo    = cd_codigo,
            usuario      = usuario,
            origem       = origem,
            auto_enviar  = auto_enviar,
        )
    except Exception as e:
        logger.error(f"[API] Erro no pipeline: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    return resultado


# ── UPLOAD EM LOTE (múltiplos arquivos) ───────────────────────────────────────

@app.post("/api/upload-lote", summary="Upload de múltiplos arquivos em lote")
async def upload_lote(
    arquivos:    List[UploadFile] = File(...),
    cd_codigo:   str              = Form(..., description="'OSA' ou 'ITJ'"),
    usuario:     str              = Form(..., description="'timoteo' ou 'carlos'"),
    origem:      str              = Form(None),
    auto_enviar: bool             = Form(True),
    orq:         Orquestrador     = Depends(get_orq),
):
    resultados = []
    for arq in arquivos:
        destino = settings.UPLOAD_DIR / f"{cd_codigo}_{usuario}_{arq.filename}"
        # Usa await arq.read() para garantir leitura completa independente do ponteiro
        conteudo = await arq.read()
        destino.write_bytes(conteudo)
        logger.info(f"[API] Lote: {arq.filename} ({len(conteudo)} bytes) | CD={cd_codigo} | user={usuario}")
        try:
            r = await orq.processar_arquivo(
                arquivo_path=destino,
                cd_codigo=cd_codigo,
                usuario=usuario,
                origem=origem,
                auto_enviar=auto_enviar,
            )
            resultados.append({"arquivo": arq.filename, "status": "ok", "resultado": r})
        except Exception as e:
            logger.error(f"[API] Erro no lote ({arq.filename}): {e}")
            resultados.append({"arquivo": arq.filename, "status": "erro", "detalhe": str(e)})

    total_ok   = sum(1 for r in resultados if r["status"] == "ok")
    total_erro = len(resultados) - total_ok
    return {"total": len(arquivos), "ok": total_ok, "erros": total_erro, "resultados": resultados}


# ── HEALTH CHECK ──────────────────────────────────────────────────────────────

@app.get("/api/ping", summary="Health check leve — não depende do banco")
async def ping():
    return {"status": "ok", "timestamp": datetime.utcnow().isoformat()}


# ── DASHBOARD ─────────────────────────────────────────────────────────────────

@app.get("/api/dashboard", summary="Dashboard consolidado tempo real")
async def dashboard(
    cd_codigo: Optional[str] = None,
    orq: Orquestrador = Depends(get_orq),
):
    return await orq.dashboard_tempo_real(cd_codigo)


# ── ONDAS DE HOJE ─────────────────────────────────────────────────────────────

@app.get("/api/ondas-hoje", summary="Ondas planejadas para um dia (default: hoje)")
async def ondas_hoje(
    cd_codigo: Optional[str] = None,
    data:      Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    data_plano = date.fromisoformat(data) if data else date.today()

    # Resolve cd_id se filtro informado
    cd_id = None
    if cd_codigo:
        res_cd = await db.execute(
            select(CentroDistribuicao).where(CentroDistribuicao.codigo == cd_codigo)
        )
        cd_obj = res_cd.scalar_one_or_none()
        if cd_obj:
            cd_id = cd_obj.id

    filtro_plano = [PlanoDia.data_plano == data_plano]
    if cd_id:
        filtro_plano.append(PlanoDia.cd_id == cd_id)

    res_planos = await db.execute(
        select(PlanoDia)
        .options(selectinload(PlanoDia.cd))
        .where(and_(*filtro_plano))
        .order_by(PlanoDia.criado_em.desc())
    )
    planos = res_planos.scalars().all()

    resultado = []
    for plano in planos:
        res_ondas = await db.execute(
            select(Onda)
            .options(selectinload(Onda.transportadora), selectinload(Onda.veiculo))
            .where(Onda.plano_id == plano.id)
            .order_by(Onda.numero_onda)
        )
        for o in res_ondas.scalars().all():
            # Conta remessas da onda
            res_count = await db.execute(
                select(func.count(OndaRemessa.remessa_id)).where(OndaRemessa.onda_id == o.id)
            )
            total_rem = res_count.scalar() or 0

            resultado.append({
                "id":              o.id,
                "nome":            o.nome,
                "tipo":            o.tipo,
                "status":          o.status,
                "transportadora":  o.transportadora.nome if o.transportadora else None,
                "veiculo":         o.veiculo.tipo if o.veiculo else None,
                "volume_m3":       float(o.volume_total_m3 or 0),
                "peso_kg":         float(o.peso_total_kg  or 0),
                "valor_nf":        float(o.valor_total_nf or 0),
                "ocupacao_pct":    float(o.ocupacao_pct   or 0),
                "horario_coleta":  o.horario_coleta.strftime("%Hh%M") if o.horario_coleta else None,
                "total_remessas":  total_rem,
                "plano_data":      str(plano.data_plano),
                "cd":              plano.cd.codigo if plano.cd else None,
                "justificativa":   o.justificativa,
            })

    aviso = None
    remessas_pendentes = 0
    if not resultado and data_plano >= date.today():
        filtro_pend = [Remessa.data_extracao == data_plano, Remessa.status == "novo"]
        if cd_id:
            filtro_pend.append(Remessa.cd_id == cd_id)
        res_pend = await db.execute(select(func.count(Remessa.id)).where(and_(*filtro_pend)))
        remessas_pendentes = res_pend.scalar() or 0
        aviso = "Nenhuma onda planejada. Faça upload do arquivo do dia para gerar o planejamento."

    return {"ondas": resultado, "aviso": aviso, "remessas_pendentes": remessas_pendentes}


# ── PRÓXIMOS DIAS (planejamento futuro) ───────────────────────────────────────

@app.get("/api/ondas/proximos-dias", summary="Resumo de planejamento dos próximos dias úteis")
async def ondas_proximos_dias(
    cd_codigo: Optional[str] = None,
    dias:      int           = 5,
    db: AsyncSession = Depends(get_db),
):
    cd_id = None
    if cd_codigo:
        res_cd = await db.execute(
            select(CentroDistribuicao).where(CentroDistribuicao.codigo == cd_codigo)
        )
        cd_obj = res_cd.scalar_one_or_none()
        if cd_obj:
            cd_id = cd_obj.id

    # Próximos N dias úteis (seg-sex), a partir de amanhã
    datas: list[date] = []
    cursor = date.today() + timedelta(days=1)
    while len(datas) < dias:
        if cursor.weekday() < 5:
            datas.append(cursor)
        cursor += timedelta(days=1)

    tem_onda = exists().where(OndaRemessa.remessa_id == Remessa.id)

    resultado = []
    for dia in datas:
        filtro_base = [Remessa.data_extracao == dia]
        if cd_id:
            filtro_base.append(Remessa.cd_id == cd_id)

        res_planejadas = await db.execute(
            select(func.count(Remessa.id)).where(and_(*filtro_base, tem_onda))
        )
        res_pendentes = await db.execute(
            select(func.count(Remessa.id)).where(
                and_(*filtro_base, Remessa.status == "novo", ~tem_onda)
            )
        )
        resultado.append({
            "data":       str(dia),
            "planejadas": res_planejadas.scalar() or 0,
            "pendentes":  res_pendentes.scalar() or 0,
        })

    return resultado


# ── HISTÓRICO DE ONDAS ────────────────────────────────────────────────────────

async def _coletar_historico_ondas(
    periodo: str, cd_codigo: Optional[str], db: AsyncSession
) -> list[dict]:
    """Lógica compartilhada entre a listagem JSON (/api/ondas/historico) e a
    exportação PDF/XLSX (/api/ondas/historico/exportar)."""
    hoje = date.today()
    if periodo == "hoje":
        data_inicio, data_fim = hoje, hoje
    elif periodo == "ontem":
        data_inicio = data_fim = hoje - timedelta(days=1)
    elif periodo == "semana":
        data_inicio, data_fim = hoje - timedelta(days=7), hoje
    elif periodo == "mes":
        data_inicio, data_fim = hoje.replace(day=1), hoje
    elif periodo == "trimestre":
        data_inicio, data_fim = hoje - timedelta(days=90), hoje
    elif periodo == "semestre":
        data_inicio, data_fim = hoje - timedelta(days=182), hoje
    elif periodo == "ano":
        data_inicio, data_fim = hoje - timedelta(days=365), hoje
    else:
        raise HTTPException(400, f"Período inválido: {periodo}")

    cd_id = None
    if cd_codigo:
        res_cd = await db.execute(select(CentroDistribuicao).where(CentroDistribuicao.codigo == cd_codigo))
        cd_obj = res_cd.scalar_one_or_none()
        if cd_obj:
            cd_id = cd_obj.id

    filtro_plano = [PlanoDia.data_plano >= data_inicio, PlanoDia.data_plano <= data_fim]
    if cd_id:
        filtro_plano.append(PlanoDia.cd_id == cd_id)

    res_planos = await db.execute(
        select(PlanoDia)
        .options(selectinload(PlanoDia.cd))
        .where(and_(*filtro_plano))
        .order_by(PlanoDia.data_plano.desc(), PlanoDia.cd_id)
    )
    planos = res_planos.scalars().all()
    if not planos:
        return []

    plano_ids = [p.id for p in planos]

    # Todas as ondas de todos os planos do período, numa única ida ao banco.
    res_ondas = await db.execute(
        select(Onda)
        .options(selectinload(Onda.transportadora))
        .where(Onda.plano_id.in_(plano_ids))
        .order_by(Onda.numero_onda)
    )
    ondas_por_plano: dict[int, list[Onda]] = {}
    onda_ids: list[int] = []
    for o in res_ondas.scalars().all():
        ondas_por_plano.setdefault(o.plano_id, []).append(o)
        onda_ids.append(o.id)

    # Contagem de remessas por onda, agregada de uma vez (era 1 query por onda).
    #
    # Nota: cheguei a rodar esta query e a de OTIF abaixo em paralelo (sessões
    # concorrentes), já que são independentes entre si. Revertido — medido contra
    # este Supabase, para o período "ano" (muitos onda_ids/plano_ids de uma vez)
    # rodar as duas ao mesmo tempo gerou contenção no lado do banco e picos de
    # até 69s, pior que as duas em série (~6s). Mantido sequencial aqui;
    # dashboard() usa paralelismo porque lá as agregações são baratas (poucas
    # linhas, índice direto), sem esse efeito.
    contagem_por_onda: dict[int, int] = {}
    if onda_ids:
        res_counts = await db.execute(
            select(OndaRemessa.onda_id, func.count(OndaRemessa.remessa_id))
            .where(OndaRemessa.onda_id.in_(onda_ids))
            .group_by(OndaRemessa.onda_id)
        )
        contagem_por_onda = dict(res_counts.all())

    # OTIF por plano (entregues/tentativas/devolvidos), agregado de uma vez
    # (era 1 query por plano) — agrupado por plano_id + status.
    otif_por_plano: dict[int, dict[str, int]] = {}
    if onda_ids:
        res_otif = await db.execute(
            select(Onda.plano_id, Remessa.status, func.count(Remessa.id))
            .join(OndaRemessa, OndaRemessa.remessa_id == Remessa.id)
            .join(Onda, Onda.id == OndaRemessa.onda_id)
            .where(Onda.plano_id.in_(plano_ids))
            .group_by(Onda.plano_id, Remessa.status)
        )
        for plano_id, status, qtd in res_otif.all():
            otif_por_plano.setdefault(plano_id, {})[status] = qtd

    resultado = []
    for plano in planos:
        ondas = ondas_por_plano.get(plano.id, [])
        if not ondas:
            continue

        ondas_out = []
        total_remessas_plano = 0
        for o in ondas:
            total_rem = contagem_por_onda.get(o.id, 0)
            total_remessas_plano += total_rem
            ondas_out.append({
                "id":             o.id,
                "numero_onda":    o.numero_onda,
                "nome":           o.nome,
                "status":         o.status,
                "transportadora": o.transportadora.nome if o.transportadora else None,
                "total_remessas": total_rem,
                "justificativa":  o.justificativa,
            })

        # OTIF do plano: entregues / (entregues + tentativas + devolvidos) entre as
        # remessas de todas as ondas deste plano.
        contagem_status = otif_por_plano.get(plano.id, {})
        entregues  = contagem_status.get("entregue", 0)
        tentativas = contagem_status.get("tentativa", 0)
        devolvidos = contagem_status.get("devolvido", 0)
        total_fin  = entregues + tentativas + devolvidos
        otif_pct   = round(entregues / total_fin * 100, 1) if total_fin > 0 else None

        # Status geral do plano, derivado do status das ondas.
        statuses_onda = {o["status"] for o in ondas_out}
        if statuses_onda == {"fechada"}:
            status_geral = "fechado"
        elif "fechada" in statuses_onda:
            status_geral = "parcial"
        else:
            status_geral = "aberto"

        resultado.append({
            "plano_id":        plano.id,
            "data_plano":      str(plano.data_plano),
            "ciclo":           plano.ciclo,
            "cd":              plano.cd.codigo if plano.cd else None,
            "cd_nome":         plano.cd.nome   if plano.cd else None,
            "total_ondas":     len(ondas_out),
            "total_remessas":  total_remessas_plano,
            "otif_pct":        otif_pct,
            "status_geral":    status_geral,
            "ondas":           ondas_out,
        })

    return resultado


@app.get("/api/ondas/historico", summary="Histórico de ondas agrupado por plano/período")
async def historico_ondas(
    periodo: str = "hoje",
    cd_codigo: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    # Cache de 120s: só vale a pena para períodos largos (mes/trimestre/semestre/
    # ano), que são os que pesam de verdade — "hoje"/"ontem" são baratos e mudam
    # o dia todo, então seguem sempre ao vivo (sem cache).
    if periodo not in ("hoje", "ontem"):
        cache_key = f"ondas_historico:{periodo}:{cd_codigo or 'todos'}"
        cached = get_cached(cache_key, ttl_seconds=120)
        if cached is not None:
            return cached
        resultado = await _coletar_historico_ondas(periodo, cd_codigo, db)
        set_cached(cache_key, resultado)
        return resultado

    return await _coletar_historico_ondas(periodo, cd_codigo, db)


PERIODO_LABELS_HISTORICO = {
    "hoje": "Hoje", "ontem": "Ontem", "semana": "Últimos 7 dias",
    "mes": "Este mês", "trimestre": "Últimos 3 meses",
    "semestre": "Últimos 6 meses", "ano": "Último ano",
}


@app.get("/api/ondas/historico/exportar", summary="Exporta o histórico de ondas em PDF ou XLSX")
async def exportar_historico_ondas(
    periodo: str = "mes",
    cd_codigo: Optional[str] = None,
    formato: str = "pdf",
    db: AsyncSession = Depends(get_db),
):
    if formato not in ("pdf", "xlsx"):
        raise HTTPException(422, "formato inválido — use 'pdf' ou 'xlsx'")

    planos = await _coletar_historico_ondas(periodo, cd_codigo, db)

    total_ondas    = sum(p["total_ondas"] for p in planos)
    total_remessas = sum(p["total_remessas"] for p in planos)
    otifs_validos  = [p["otif_pct"] for p in planos if p["otif_pct"] is not None]
    otif_medio     = round(sum(otifs_validos) / len(otifs_validos), 1) if otifs_validos else None

    periodo_label = PERIODO_LABELS_HISTORICO.get(periodo, periodo)
    nome_arquivo  = f"historico_ondas_{periodo}_{date.today():%Y%m%d}"

    if formato == "xlsx":
        return _historico_ondas_xlsx(
            planos, periodo_label, cd_codigo, total_ondas, total_remessas, otif_medio, nome_arquivo,
        )
    return _historico_ondas_pdf(
        planos, periodo_label, cd_codigo, total_ondas, total_remessas, otif_medio, nome_arquivo,
    )


def _historico_ondas_pdf(
    planos, periodo_label, cd_codigo, total_ondas, total_remessas, otif_medio, nome_arquivo,
):
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
    )

    BD_BLUE   = colors.HexColor("#1A3F6F")
    GRAY_ROW  = colors.HexColor("#F5F7FA")
    GRAY_LINE = colors.HexColor("#E5E7EB")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.8*cm, bottomMargin=1.8*cm,
    )
    styles = getSampleStyleSheet()

    def st(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    title_st = st("T", textColor=BD_BLUE, fontSize=16, fontName="Helvetica-Bold", spaceAfter=2)
    sub_st   = st("S", textColor=colors.HexColor("#6B7280"), fontSize=9, spaceAfter=6)
    cell_st  = st("C", fontSize=7, leading=9)
    foot_st  = st("F", fontSize=6, textColor=colors.HexColor("#9CA3AF"), alignment=TA_CENTER)

    cd_txt = cd_codigo or "Todos os CDs"
    story = [
        Paragraph("Histórico de Ondas", title_st),
        Paragraph(
            f"Período: {periodo_label} · CD: {cd_txt} · "
            f"Gerado em {datetime.now():%d/%m/%Y %Hh%M}",
            sub_st,
        ),
        HRFlowable(width="100%", thickness=0.5, color=GRAY_LINE),
        Spacer(1, 0.3*cm),
    ]

    otif_txt = f"{otif_medio:.1f}%" if otif_medio is not None else "sem dados"
    resumo_rows = [["Planos", str(len(planos)), "Ondas", str(total_ondas),
                     "Remessas", str(total_remessas), "OTIF médio", otif_txt]]
    t_resumo = Table(resumo_rows, colWidths=[2*cm, 1.6*cm, 2*cm, 1.6*cm, 2.2*cm, 1.6*cm, 2.4*cm, 1.6*cm])
    t_resumo.setStyle(TableStyle([
        ("FONTNAME",      (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 8),
        ("TEXTCOLOR",     (0, 0), (-1, -1), colors.HexColor("#4B5563")),
        ("GRID",          (0, 0), (-1, -1), 0.4, GRAY_LINE),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("BACKGROUND",    (0, 0), (0, 0), GRAY_ROW),
        ("BACKGROUND",    (2, 0), (2, 0), GRAY_ROW),
        ("BACKGROUND",    (4, 0), (4, 0), GRAY_ROW),
        ("BACKGROUND",    (6, 0), (6, 0), GRAY_ROW),
    ]))
    story.append(t_resumo)
    story.append(Spacer(1, 0.4*cm))

    header = ["Data", "CD", "Ciclo", "Onda", "Transportadora", "Status Onda", "Remessas", "OTIF Plano", "Status Plano"]
    rows = [header]
    for plano in planos:
        otif_plano_txt = f"{plano['otif_pct']:.1f}%" if plano["otif_pct"] is not None else "—"
        for onda in plano["ondas"]:
            rows.append([
                plano["data_plano"], plano["cd"] or "—", str(plano["ciclo"]),
                onda["nome"], onda["transportadora"] or "—", onda["status"],
                str(onda["total_remessas"]), otif_plano_txt, plano["status_geral"],
            ])

    if len(rows) > 1:
        t = Table(
            rows,
            colWidths=[2.2*cm, 1.3*cm, 1.3*cm, 5.5*cm, 4*cm, 2.3*cm, 2*cm, 2.2*cm, 2.2*cm],
            repeatRows=1,
        )
        t.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0), BD_BLUE),
            ("TEXTCOLOR",      (0, 0), (-1, 0), colors.white),
            ("FONTNAME",       (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",       (0, 0), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRAY_ROW]),
            ("GRID",           (0, 0), (-1, -1), 0.4, GRAY_LINE),
            ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",     (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 4),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("Nenhuma onda encontrada no período.", cell_st))

    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.4, color=GRAY_LINE))
    story.append(Paragraph("Becton Dickinson — PEO-BD v1.0  |  Documento gerado automaticamente", foot_st))

    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={nome_arquivo}.pdf"},
    )


def _historico_ondas_xlsx(
    planos, periodo_label, cd_codigo, total_ondas, total_remessas, otif_medio, nome_arquivo,
):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Histórico de Ondas"

    azul_bd     = PatternFill("solid", fgColor="1A3F6F")
    header_font = Font(bold=True, color="FFFFFF")
    bold        = Font(bold=True)

    cd_txt = cd_codigo or "Todos os CDs"
    ws["A1"] = "Histórico de Ondas"
    ws["A1"].font = Font(bold=True, size=14, color="1A3F6F")
    ws["A2"] = f"Período: {periodo_label} · CD: {cd_txt}"
    ws["A2"].font = Font(italic=True, color="6B7280")
    ws["A3"] = f"Gerado em {datetime.now():%d/%m/%Y %Hh%M}"
    ws["A3"].font = Font(italic=True, size=8, color="9CA3AF")

    resumo = [
        ("Planos",          len(planos)),
        ("Ondas",           total_ondas),
        ("Remessas",        total_remessas),
        ("OTIF médio (%)",  otif_medio if otif_medio is not None else "sem dados"),
    ]
    linha = 5
    for label, valor in resumo:
        ws.cell(row=linha, column=1, value=label).font = bold
        ws.cell(row=linha, column=2, value=valor)
        linha += 1

    linha += 1
    header = ["Data", "CD", "Ciclo", "Onda", "Transportadora", "Status Onda", "Remessas", "OTIF Plano (%)", "Status Plano"]
    for col, texto in enumerate(header, start=1):
        c = ws.cell(row=linha, column=col, value=texto)
        c.font = header_font
        c.fill = azul_bd
    linha += 1
    for plano in planos:
        for onda in plano["ondas"]:
            ws.append([
                plano["data_plano"], plano["cd"] or "—", plano["ciclo"],
                onda["nome"], onda["transportadora"] or "—", onda["status"],
                onda["total_remessas"], plano["otif_pct"], plano["status_geral"],
            ])

    larguras = [12, 8, 7, 32, 24, 14, 11, 14, 13]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome_arquivo}.xlsx"},
    )


@app.get("/api/ondas/{onda_id}/remessas", summary="Lista as remessas de uma onda")
async def remessas_da_onda(onda_id: int, db: AsyncSession = Depends(get_db)):
    res_onda = await db.execute(
        select(Onda).options(selectinload(Onda.transportadora)).where(Onda.id == onda_id)
    )
    onda = res_onda.scalar_one_or_none()
    if not onda:
        raise HTTPException(404, "Onda não encontrada")

    res = await db.execute(
        select(Remessa)
        .options(selectinload(Remessa.cliente))
        .join(OndaRemessa, OndaRemessa.remessa_id == Remessa.id)
        .where(OndaRemessa.onda_id == onda_id)
        .order_by(OndaRemessa.sequencia)
    )
    remessas = res.scalars().all()

    return {
        "onda": {
            "id":             onda.id,
            "numero_onda":    onda.numero_onda,
            "nome":           onda.nome,
            "status":         onda.status,
            "transportadora": onda.transportadora.nome if onda.transportadora else None,
            "justificativa":  onda.justificativa,
        },
        "remessas": [
            {
                "id":             r.id,
                "numero_remessa": r.numero_remessa,
                "cliente":        r.cliente.razao_social if r.cliente else "—",
                "cidade_uf":      f"{r.cliente.cidade or ''}/{r.cliente.uf or ''}".strip("/") if r.cliente else "—",
                "status":         r.status,
                "volume_m3":      float(r.volume_m3 or 0),
                "peso_kg":        float(r.peso_kg or 0),
                "valor_nf":       float(r.valor_nf or 0),
                "prioridade":     r.prioridade,
            }
            for r in remessas
        ],
    }


class FechamentoOndaSchema(BaseModel):
    resultados: list[dict]   # [{"remessa_id": 1, "status": "entregue"}, ...]
    usuario: Optional[str] = None


@app.post("/api/ondas/{onda_id}/fechar", summary="Registra o retorno do romaneio e fecha a onda")
async def fechar_onda(
    onda_id: int,
    payload: FechamentoOndaSchema,
    db: AsyncSession = Depends(get_db),
):
    from core.historico import HistoricoService
    historico = HistoricoService(db)

    STATUS_VALIDOS = {"entregue", "tentativa", "devolvido"}

    res_onda = await db.execute(
        select(Onda).options(selectinload(Onda.plano)).where(Onda.id == onda_id)
    )
    onda = res_onda.scalar_one_or_none()
    if not onda:
        raise HTTPException(404, "Onda não encontrada")
    if onda.status == "fechada":
        raise HTTPException(400, "Onda já está fechada")

    contagem_resultado: dict[str, int] = {}
    atualizados = 0
    for item in payload.resultados:
        remessa_id = item.get("remessa_id")
        novo_status = item.get("status")
        if novo_status not in STATUS_VALIDOS:
            raise HTTPException(400, f"Status inválido para fechamento: {novo_status}")

        res_rem = await db.execute(select(Remessa).where(Remessa.id == remessa_id))
        remessa = res_rem.scalar_one_or_none()
        if not remessa:
            continue

        remessa.status = novo_status
        atualizados += 1
        contagem_resultado[novo_status] = contagem_resultado.get(novo_status, 0) + 1

    onda.status = "fechada"

    resumo = ", ".join(f"{qtd} {status}" for status, qtd in contagem_resultado.items())
    ator_nome = payload.usuario or "operador"
    await historico.registrar(
        tipo_evento  = "mudanca_status",
        origem       = "usuario",
        ator_tipo    = "usuario_humano",
        ator_nome    = ator_nome,
        cd_id        = onda.plano.cd_id if onda.plano else None,
        descricao    = (
            f"{onda.nome or f'Onda {onda.numero_onda:02d}'} fechada por retorno de romaneio "
            f"({atualizados} remessa(s): {resumo})"
        ),
        resultado    = "sucesso",
        gravidade    = None,
        dados_extra  = {
            "onda_id": onda.id,
            "atualizados": atualizados,
            "resultado_por_status": contagem_resultado,
        },
    )

    await db.commit()

    return {
        "status":       "fechada",
        "onda_id":      onda.id,
        "atualizados":  atualizados,
        "resultado_por_status": contagem_resultado,
    }


# ── ALERTAS ───────────────────────────────────────────────────────────────────

@app.get("/api/alertas", summary="Lista alertas não resolvidos")
async def listar_alertas(
    cd_id:      Optional[int] = None,
    severidade: Optional[str] = None,
    db:         AsyncSession  = Depends(get_db),
):
    filtros = [Alerta.resolvido == False]
    if cd_id:
        filtros.append(Alerta.cd_id == cd_id)
    if severidade:
        filtros.append(Alerta.severidade == severidade)

    res = await db.execute(
        select(Alerta)
        .where(and_(*filtros))
        .order_by(Alerta.criado_em.desc())
        .limit(100)
    )
    alertas = res.scalars().all()

    return [
        {
            "id":         a.id,
            "tipo":       a.tipo,
            "severidade": a.severidade,
            "titulo":     a.titulo,
            "descricao":  a.descricao,
            "remessa_id": a.remessa_id,
            "criado_em":  a.criado_em.isoformat() if a.criado_em else None,
        }
        for a in alertas
    ]


@app.post("/api/alertas/{alerta_id}/resolver", summary="Resolve alerta e registra no histórico")
async def resolver_alerta(alerta_id: int, db: AsyncSession = Depends(get_db)):
    from datetime import datetime
    from core.historico import HistoricoService

    res = await db.execute(select(Alerta).where(Alerta.id == alerta_id))
    alerta = res.scalar_one_or_none()
    if not alerta:
        raise HTTPException(404, "Alerta não encontrado")
    if alerta.resolvido:
        return {"status": "resolvido"}

    alerta.resolvido    = True
    alerta.resolvido_em = datetime.utcnow()

    historico = HistoricoService(db)
    await historico.registrar(
        tipo_evento  = "mudanca_status",
        origem       = "usuario",
        ator_tipo    = "usuario_humano",
        ator_nome    = "operador",
        remessa_id   = alerta.remessa_id,
        cd_id        = alerta.cd_id,
        descricao    = "Alerta resolvido manualmente pelo operador",
        resultado    = "sucesso",
        gravidade    = alerta.severidade,
        dados_extra  = {"alerta_id": alerta.id, "tipo_alerta": alerta.tipo},
    )

    await db.commit()
    return {"status": "resolvido"}


# ── REMESSAS ──────────────────────────────────────────────────────────────────

class RemessaManualSchema(BaseModel):
    numero_remessa:      str
    cliente_nome:        str
    cidade:              str | None  = None
    volume_m3:           float       = 0.0
    peso_kg:             float       = 0.0
    transportadora_nome: str | None  = None
    origem_contingencia: bool        = True


@app.post("/api/remessas/manual", summary="Recebe remessa criada durante contingência offline")
async def criar_remessa_manual(
    payload: RemessaManualSchema,
    db: AsyncSession = Depends(get_db),
):
    """
    Endpoint de sincronização do Modo de Contingência.
    Aceita remessas registradas offline no IndexedDB e as persiste no banco.
    Aplica o mesmo hash de deduplicação do AgenteIngestor para evitar duplicatas
    em caso de reconexão parcial ou múltiplas tentativas de sync.
    """
    from core.historico import HistoricoService
    historico = HistoricoService(db)

    # ── Hash de deduplicação (mesma lógica do AgenteIngestor) ─────────────────
    dados_hash = "|".join([
        payload.numero_remessa.strip(),
        str(payload.peso_kg),
        str(payload.volume_m3),
        "",   # valor_nf — não disponível em contingência
        "",   # numero_nf — não disponível em contingência
    ])
    hash_val = hashlib.sha256(dados_hash.encode()).hexdigest()

    # Verifica duplicata por hash ou número
    res_hash = await db.execute(select(Remessa).where(Remessa.hash_remessa == hash_val))
    if res_hash.scalar_one_or_none():
        return {"status": "duplicata", "numero_remessa": payload.numero_remessa, "sincronizado": True}

    res_num = await db.execute(select(Remessa).where(Remessa.numero_remessa == payload.numero_remessa))
    if res_num.scalar_one_or_none():
        return {"status": "duplicata", "numero_remessa": payload.numero_remessa, "sincronizado": True}

    # ── Resolve ou cria cliente ───────────────────────────────────────────────
    cliente = None
    if payload.cliente_nome:
        res_cli = await db.execute(
            select(Cliente).where(Cliente.razao_social.ilike(f"%{payload.cliente_nome[:30]}%"))
        )
        cliente = res_cli.scalar_one_or_none()
        if not cliente:
            cliente = Cliente(
                razao_social    = payload.cliente_nome,
                tipo            = "outros",
                cidade          = payload.cidade,
                tem_armazenagem = False,
                janela_flexivel = True,
            )
            db.add(cliente)
            await db.flush()

    # ── Resolve CD padrão (OSA) ───────────────────────────────────────────────
    res_cd = await db.execute(
        select(CentroDistribuicao).where(CentroDistribuicao.ativo == True).limit(1)
    )
    cd = res_cd.scalar_one_or_none()

    # ── Cria remessa ──────────────────────────────────────────────────────────
    remessa = Remessa(
        numero_remessa = payload.numero_remessa.strip(),
        origem         = "manual_contingencia",
        cd_id          = cd.id if cd else None,
        cliente_id     = cliente.id if cliente else None,
        data_extracao  = date.today(),
        volume_m3      = payload.volume_m3 or None,
        peso_kg        = payload.peso_kg or None,
        status         = "novo",
        prioridade     = "normal",
        hash_remessa   = hash_val,
        nf_emitida     = False,
    )
    db.add(remessa)
    await db.flush()

    # ── Registra no histórico ──────────────────────────────────────────────────
    await historico.registrar(
        tipo_evento  = "upload_processado",
        origem       = "manual_contingencia",
        ator_tipo    = "usuario_humano",
        ator_nome    = "Operador (contingência)",
        remessa_id   = remessa.id,
        cd_id        = cd.id if cd else None,
        descricao    = (
            f"Remessa {payload.numero_remessa} registrada manualmente durante "
            f"modo de contingência — transportadora: {payload.transportadora_nome or 'N/I'}"
        ),
        resultado    = "sucesso",
        gravidade    = None,
        dados_extra  = {
            "origem_contingencia": True,
            "transportadora_nome": payload.transportadora_nome,
            "cidade": payload.cidade,
        },
    )

    await db.commit()

    return {
        "status":          "criado",
        "numero_remessa":  payload.numero_remessa,
        "remessa_id":      remessa.id,
        "sincronizado":    True,
    }


@app.get("/api/remessas", summary="Lista remessas")
async def listar_remessas(
    cd_id:             Optional[int]  = None,
    data:              Optional[date] = None,
    status:            Optional[str]  = None,
    prioridade:        Optional[str]  = None,
    transportadora_id: Optional[int]  = None,
    sem_onda:          Optional[bool] = None,
    nf_emitida:        Optional[bool] = None,
    limit:             int            = 500,
    db: AsyncSession = Depends(get_db),
):
    filtros = []
    if cd_id:
        filtros.append(Remessa.cd_id == cd_id)
    if data:
        filtros.append(Remessa.data_extracao == data)
    if status:
        filtros.append(Remessa.status == status)
    if prioridade:
        filtros.append(Remessa.prioridade == prioridade)
    if nf_emitida is not None:
        filtros.append(Remessa.nf_emitida == nf_emitida)
    if sem_onda:
        # Nunca foi agrupada em nenhuma onda — mesma definição usada em
        # /api/admin/reprocessar-historico e no card "Pendentes" do dashboard.
        filtros.append(~exists().where(OndaRemessa.remessa_id == Remessa.id))

    query = (
        select(Remessa)
        # joinedload (não selectinload) aqui de propósito: Remessa.cliente é
        # many-to-one (1 cliente por remessa) — o LEFT JOIN não duplica linhas
        # de Remessa, então é seguro e colapsa 2 round-trips de rede (query base
        # + query separada do selectinload) em 1 só. Contra o Postgres remoto
        # (Supabase), isso é metade da latência fixa por request.
        .options(joinedload(Remessa.cliente))
        .order_by(Remessa.prioridade, Remessa.janela_inicio)
    )
    if transportadora_id == 0:
        # Sentinela "sem transportadora definida": remessas cuja onda ainda não
        # tem transportadora resolvida (Comunicador não rodou / não encontrou
        # a transportadora), ou que ainda não foram montadas em nenhuma onda.
        subq_com_transportadora = (
            select(OndaRemessa.remessa_id)
            .join(Onda, Onda.id == OndaRemessa.onda_id)
            .where(Onda.transportadora_id.isnot(None))
        )
        query = query.where(Remessa.id.not_in(subq_com_transportadora))
    elif transportadora_id:
        # Remessa não tem transportadora_id direto — o vínculo é via a onda
        # em que ela foi montada (onda_remessas → ondas.transportadora_id).
        query = (
            query
            .join(OndaRemessa, OndaRemessa.remessa_id == Remessa.id)
            .join(Onda, Onda.id == OndaRemessa.onda_id)
            .where(Onda.transportadora_id == transportadora_id)
            .distinct()
        )
    if filtros:
        query = query.where(and_(*filtros))

    res      = await db.execute(query.limit(limit))
    remessas = res.scalars().all()

    return [
        {
            "id":             r.id,
            "numero":         r.numero_remessa,
            "status":         r.status,
            "prioridade":     r.prioridade,
            "is_ata":         r.is_ata,
            "dias_restantes": r.dias_restantes,
            "nf_emitida":     r.nf_emitida,
            "numero_empenho": r.numero_empenho,
            "motivo_tentativa": r.motivo_tentativa,
            "volume_m3":      float(r.volume_m3 or 0),
            "peso_kg":        float(r.peso_kg   or 0),
            "valor_nf":       float(r.valor_nf  or 0),
            "janela":         (
                f"{r.janela_inicio.strftime('%Hh%M')} – {r.janela_fim.strftime('%Hh%M')}"
                if r.janela_inicio and r.janela_fim else "Flexível"
            ),
            "cliente":    r.cliente.razao_social if r.cliente else None,
            "cidade":     r.cliente.cidade if r.cliente else None,
            "uf":         r.cliente.uf if r.cliente else None,
            "cidade_uf":  f"{r.cliente.cidade or ''}/{r.cliente.uf or ''}" if r.cliente else None,
        }
        for r in remessas
    ]


# ── OPORTUNIDADES DE CONSOLIDAÇÃO FTL ─────────────────────────────────────────

@app.get("/api/oportunidades", summary="Oportunidades de consolidação FTL")
async def listar_oportunidades(
    status: Optional[str] = "aberta",
    db: AsyncSession = Depends(get_db),
):
    # Default "aberta" — sem isso, oportunidades "descartada" (duplicatas
    # substituídas por uma mais recente da mesma região, ver
    # AgenteClassificador._analisar_consolidacao) afogam as ativas na
    # listagem. ?status=todas remove o filtro para ver o histórico completo.
    filtros = []
    if status and status != "todas":
        filtros.append(OportunidadeConsolidacao.status == status)

    query = select(OportunidadeConsolidacao).order_by(
        OportunidadeConsolidacao.economia_estimada.desc()
    )
    if filtros:
        query = query.where(and_(*filtros))

    res   = await db.execute(query.limit(50))
    opors = res.scalars().all()

    resultado = []
    for o in opors:
        # A oportunidade não guarda o vínculo com remessas específicas — foi
        # calculada por agregação (cd + região). Reconstituímos aqui as
        # remessas que hoje ainda compõem esse agrupamento (fracionado,
        # cliente com armazenagem, ainda em aberto) para exibir o detalhe.
        res_rem = await db.execute(
            select(Remessa)
            .options(selectinload(Remessa.cliente))
            .join(Cliente, Remessa.cliente_id == Cliente.id)
            .where(
                and_(
                    Remessa.cd_id       == o.cd_id,
                    Remessa.status.in_(["novo", "planejado", "coletado"]),
                    Remessa.tipo_entrega == "fracionado",
                    Cliente.tem_armazenagem == True,
                    Cliente.regiao      == o.regiao,
                )
            )
            .order_by(Remessa.criado_em.desc())
            .limit(50)
        )
        remessas_op = res_rem.scalars().all()

        resultado.append({
            "id":                 o.id,
            "cd_id":              o.cd_id,
            "regiao":             o.regiao,
            "data_analise":       str(o.data_analise) if o.data_analise else None,
            "qtd_clientes":       o.qtd_clientes,
            "volume_atual_m3":    float(o.volume_atual_m3 or 0),
            "tipo_atual":         o.tipo_atual,
            "tipo_possivel":      o.tipo_possivel,
            "economia_estimada":  float(o.economia_estimada or 0),
            "acao_sugerida":      o.acao_sugerida,
            "status":             o.status,
            "criado_em":          o.criado_em.isoformat() if o.criado_em else None,
            "clientes":           sorted({r.cliente.razao_social for r in remessas_op if r.cliente}),
            "valor_nf_total":     float(sum(r.valor_nf or 0 for r in remessas_op)),
            "remessas": [
                {
                    "id":         r.id,
                    "numero":     r.numero_remessa,
                    "cliente":    r.cliente.razao_social if r.cliente else "—",
                    "volume_m3":  float(r.volume_m3 or 0),
                    "peso_kg":    float(r.peso_kg or 0),
                    "valor_nf":   float(r.valor_nf or 0),
                }
                for r in remessas_op
            ],
        })

    return resultado


@app.post("/api/oportunidades/analisar", summary="Varre remessas recentes em busca de oportunidades de consolidação FTL")
async def analisar_oportunidades(
    dias: int = 30,
    db: AsyncSession = Depends(get_db),
):
    """
    Roda sob demanda (botão "Recalcular" na página Oportunidades FTL).
    Ao contrário da análise automática de classificar_upload()/
    classificar_periodo() — que só olha o upload ou dia do momento —
    esta varredura agrega os últimos `dias` dias de remessas ainda em
    aberto, agrupando por região, para achar padrões de consolidação que só
    aparecem olhando várias datas juntas.
    """
    classificador = AgenteClassificador(db)
    return await classificador.analisar_oportunidades_periodo(dias)


# ── PLANOS E ONDAS ────────────────────────────────────────────────────────────

@app.get("/api/planos/{plano_id}", summary="Detalhe do plano e ondas")
async def detalhe_plano(plano_id: int, db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(PlanoDia).where(PlanoDia.id == plano_id))
    plano = res.scalar_one_or_none()
    if not plano:
        raise HTTPException(404, "Plano não encontrado")

    res_ondas = await db.execute(
        select(Onda).where(Onda.plano_id == plano_id).order_by(Onda.numero_onda)
    )
    ondas = res_ondas.scalars().all()

    return {
        "plano": {
            "id":             plano.id,
            "data":           str(plano.data_plano),
            "ciclo":          plano.ciclo,
            "status":         plano.status,
            "total_remessas": plano.total_remessas,
            "volume_m3":      float(plano.total_volume_m3 or 0),
            "valor_nf":       float(plano.total_valor_nf  or 0),
        },
        "ondas": [
            {
                "id":            o.id,
                "numero":        o.numero_onda,
                "nome":          o.nome,
                "tipo":          o.tipo,
                "status":        o.status,
                "volume_m3":     float(o.volume_total_m3 or 0),
                "peso_kg":       float(o.peso_total_kg   or 0),
                "valor_nf":      float(o.valor_total_nf  or 0),
                "ocupacao_pct":  float(o.ocupacao_pct    or 0),
                "transportadora": o.transportadora.nome if o.transportadora else None,
                "veiculo":        o.veiculo.tipo         if o.veiculo       else None,
                "horario_coleta": str(o.horario_coleta)  if o.horario_coleta else None,
                "justificativa":  o.justificativa,
            }
            for o in ondas
        ],
    }


@app.patch("/api/planos/{plano_id}/aprovar", summary="Aprova plano e libera envio")
async def aprovar_plano(
    plano_id: int,
    usuario:  str = Form(...),
    db: AsyncSession = Depends(get_db),
    orq: Orquestrador = Depends(get_orq),
):
    from datetime import datetime
    res = await db.execute(select(PlanoDia).where(PlanoDia.id == plano_id))
    plano = res.scalar_one_or_none()
    if not plano:
        raise HTTPException(404, "Plano não encontrado")
    plano.status       = "aprovado"
    plano.aprovado_por = usuario
    plano.aprovado_em  = datetime.utcnow()
    await db.commit()

    rel = await orq.comunicador.programar_coletas(plano_id)
    return {"status": "aprovado", "comunicacao": rel}


# ── RASTREIO ──────────────────────────────────────────────────────────────────

@app.post("/api/rastreio/update", summary="Registra update de rastreio (motorista/manual)")
async def rastreio_update(
    numero_remessa: str = Form(...),
    status:         str = Form(...),
    transportadora: str = Form(...),
    localizacao:    str = Form(None),
    fonte:          str = Form("manual"),
    orq: Orquestrador = Depends(get_orq),
):
    return await orq.registrar_update_rastreio(
        numero_remessa, status, transportadora, localizacao, fonte
    )


@app.post("/api/rastreio/sla-check", summary="Verifica SLA e gera alertas")
async def sla_check(orq: Orquestrador = Depends(get_orq)):
    return await orq.verificar_sla_e_alertas()


# ── DOWNLOAD DE PLANILHA ──────────────────────────────────────────────────────

@app.get("/api/outputs/{filename}", summary="Download de planilha gerada")
async def download_output(filename: str):
    for f in settings.OUTPUT_DIR.rglob(filename):
        return FileResponse(f, filename=filename)
    raise HTTPException(404, "Arquivo não encontrado")


# ── CLIENTES ─────────────────────────────────────────────────────────────────

@app.get("/api/clientes", summary="Lista clientes com estatísticas agregadas")
async def listar_clientes(
    busca: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    query = select(Cliente).order_by(Cliente.razao_social)
    if busca:
        query = query.where(Cliente.razao_social.ilike(f"%{busca}%"))

    res = await db.execute(query)
    clientes = res.scalars().all()

    resultado = []
    for c in clientes:
        # Agrega remessas
        res_rem = await db.execute(
            select(
                func.count(Remessa.id).label("total"),
                func.sum(Remessa.valor_nf).label("valor_total"),
                func.sum(Remessa.volume_m3).label("volume_total"),
            ).where(Remessa.cliente_id == c.id)
        )
        agg = res_rem.one()

        res_ent = await db.execute(
            select(func.count(Remessa.id)).where(
                and_(Remessa.cliente_id == c.id, Remessa.status == "entregue")
            )
        )
        qtd_entregues = res_ent.scalar() or 0

        res_pend = await db.execute(
            select(func.count(Remessa.id)).where(
                and_(
                    Remessa.cliente_id == c.id,
                    Remessa.status.not_in(["entregue", "devolvido"]),
                )
            )
        )
        qtd_pendentes = res_pend.scalar() or 0

        res_alertas = await db.execute(
            select(func.count(Alerta.id)).where(
                and_(Alerta.cliente_id == c.id, Alerta.resolvido == False)
            )
        )
        qtd_alertas = res_alertas.scalar() or 0

        vol = float(agg.volume_total or 0)
        resultado.append({
            "id":                 c.id,
            "razao_social":       c.razao_social,
            "tipo":               c.tipo,
            "cidade":             c.cidade,
            "uf":                 c.uf,
            "regiao":             c.regiao,
            "tem_armazenagem":    c.tem_armazenagem,
            "janela_flexivel":    c.janela_flexivel,
            "contrato_ata":       c.contrato_ata,
            "total_remessas":     agg.total or 0,
            "valor_total":        float(agg.valor_total or 0),
            "volume_acumulado_m3": round(vol, 2),
            "qtd_entregues":      qtd_entregues,
            "qtd_pendentes":      qtd_pendentes,
            "qtd_alertas":        qtd_alertas,
            "ftl_possivel":       vol >= settings.LIMIAR_FTL_M3,
        })

    return resultado


@app.get("/api/clientes/{cliente_id}", summary="Detalhe do cliente")
async def detalhe_cliente(cliente_id: int, db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(Cliente).where(Cliente.id == cliente_id))
    c = res.scalar_one_or_none()
    if not c:
        raise HTTPException(404, "Cliente não encontrado")

    res_rem = await db.execute(
        select(Remessa).where(Remessa.cliente_id == c.id).order_by(Remessa.criado_em.desc()).limit(50)
    )
    remessas = res_rem.scalars().all()

    status_count: dict = {}
    for r in remessas:
        status_count[r.status] = status_count.get(r.status, 0) + 1

    total_val  = sum(float(r.valor_nf  or 0) for r in remessas)
    total_vol  = sum(float(r.volume_m3 or 0) for r in remessas)
    entregues  = status_count.get("entregue", 0)
    pendentes  = sum(v for k, v in status_count.items() if k not in ("entregue", "devolvido"))

    res_al = await db.execute(
        select(Alerta)
        .where(and_(Alerta.cliente_id == c.id, Alerta.resolvido == False))
        .order_by(Alerta.criado_em.desc())
        .limit(20)
    )
    alertas = res_al.scalars().all()

    return {
        "id":              c.id,
        "razao_social":    c.razao_social,
        "cnpj":            c.cnpj,
        "tipo":            c.tipo,
        "cidade":          c.cidade,
        "uf":              c.uf,
        "regiao":          c.regiao,
        "cep":             c.cep,
        "tem_armazenagem": c.tem_armazenagem,
        "janela_inicio":   c.janela_inicio.strftime("%Hh%M") if c.janela_inicio else None,
        "janela_fim":      c.janela_fim.strftime("%Hh%M")    if c.janela_fim    else None,
        "janela_flexivel": c.janela_flexivel,
        "contrato_ata":    c.contrato_ata,
        "prazo_ata_dias":  c.prazo_ata_dias,
        "volume_medio_m3": float(c.volume_medio_m3 or 0),
        "resumo_remessas": {
            "total":     len(remessas),
            "entregues": entregues,
            "pendentes": pendentes,
            "por_status": status_count,
        },
        "valor_total":        round(total_val, 2),
        "volume_acumulado_m3": round(total_vol, 2),
        "ftl_possivel":       total_vol >= settings.LIMIAR_FTL_M3,
        "alertas": [
            {
                "id":         a.id,
                "tipo":       a.tipo,
                "severidade": a.severidade,
                "titulo":     a.titulo,
                "descricao":  a.descricao,
                "criado_em":  a.criado_em.isoformat() if a.criado_em else None,
            }
            for a in alertas
        ],
    }


# ── ERROS DE UPLOAD ───────────────────────────────────────────────────────────

@app.get("/api/erros-upload", summary="Painel de Diagnóstico: duplicatas, abertos, sem NF (rota mantida por compatibilidade)")
async def erros_upload(
    cd_codigo: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    cd_id = None
    if cd_codigo:
        r = await db.execute(select(CentroDistribuicao).where(CentroDistribuicao.codigo == cd_codigo))
        cd = r.scalar_one_or_none()
        if cd:
            cd_id = cd.id

    # Último upload
    q_up = select(Upload).order_by(Upload.criado_em.desc()).limit(1)
    if cd_id:
        q_up = q_up.where(Upload.cd_id == cd_id)
    res_up = await db.execute(q_up)
    ultimo_upload = res_up.scalar_one_or_none()

    # 1. Remessas duplicadas (com hash duplicado no último upload)
    duplicatas = []
    if ultimo_upload:
        res_dup = await db.execute(
            select(Remessa)
            .options(selectinload(Remessa.cliente))
            .where(
                and_(
                    Remessa.upload_id == ultimo_upload.id,
                    Remessa.duplicata_de.isnot(None),
                )
            )
            .limit(50)
        )
        for r in res_dup.scalars().all():
            duplicatas.append({
                "numero":    r.numero_remessa,
                "cliente":   r.cliente.razao_social if r.cliente else "—",
                "valor_nf":  float(r.valor_nf or 0),
                "motivo":    "Hash idêntico a remessa anterior no banco",
            })

    # 2. Remessas em Aberto (backlog): status novo e data_extracao < hoje
    hoje = date.today()
    q_ab = (
        select(Remessa)
        .options(selectinload(Remessa.cliente))
        .where(and_(Remessa.status == "novo", Remessa.data_extracao < hoje))
        .order_by(Remessa.data_extracao)
        .limit(100)
    )
    if cd_id:
        q_ab = q_ab.where(Remessa.cd_id == cd_id)
    res_ab = await db.execute(q_ab)
    em_aberto = []
    for r in res_ab.scalars().all():
        dias = (hoje - r.data_extracao).days if r.data_extracao else 0
        em_aberto.append({
            "numero":          r.numero_remessa,
            "cliente":         r.cliente.razao_social if r.cliente else "—",
            "data_extracao":   str(r.data_extracao),
            "dias_em_aberto":  dias,
            "valor_nf":        float(r.valor_nf or 0),
            "is_ata":          r.is_ata,
            "prioridade":      r.prioridade,
        })

    # 3. Remessas sem NF
    q_nf = (
        select(Remessa)
        .options(selectinload(Remessa.cliente))
        .where(
            and_(
                Remessa.nf_emitida == False,
                Remessa.status.not_in(["entregue", "devolvido"]),
            )
        )
        .order_by(Remessa.prioridade, Remessa.criado_em.desc())
        .limit(100)
    )
    if cd_id:
        q_nf = q_nf.where(Remessa.cd_id == cd_id)
    res_nf = await db.execute(q_nf)
    sem_nf = []
    for r in res_nf.scalars().all():
        sem_nf.append({
            "numero":    r.numero_remessa,
            "cliente":   r.cliente.razao_social if r.cliente else "—",
            "status":    r.status,
            "is_ata":    r.is_ata,
            "prioridade": r.prioridade,
            "valor_nf":  float(r.valor_nf or 0),
        })

    return {
        "ultimo_upload": {
            "id":         ultimo_upload.id          if ultimo_upload else None,
            "arquivo":    ultimo_upload.arquivo_nome if ultimo_upload else None,
            "criado_em":  ultimo_upload.criado_em.isoformat() if ultimo_upload else None,
        },
        "totais": {
            "duplicatas": len(duplicatas),
            "em_aberto":  len(em_aberto),
            "sem_nf":     len(sem_nf),
        },
        "duplicatas": duplicatas,
        "em_aberto":  em_aberto,
        "sem_nf":     sem_nf,
    }


@app.get("/api/uploads/{upload_id}", summary="Detalhe de um upload e as remessas associadas")
async def detalhe_upload(upload_id: int, db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(Upload).options(selectinload(Upload.cd)).where(Upload.id == upload_id)
    )
    upload = res.scalar_one_or_none()
    if not upload:
        raise HTTPException(404, "Upload não encontrado")

    res_rem = await db.execute(
        select(Remessa)
        .options(selectinload(Remessa.cliente))
        .where(Remessa.upload_id == upload_id)
        .order_by(Remessa.criado_em.desc())
    )
    remessas = res_rem.scalars().all()

    return {
        "id":             upload.id,
        "arquivo":        upload.arquivo_nome,
        "cd":             upload.cd.codigo if upload.cd else None,
        "usuario":        upload.usuario,
        "formato":        upload.formato,
        "status":         upload.status,
        "total_linhas":   upload.total_linhas,
        "linhas_validas": upload.linhas_validas,
        "linhas_erro":    upload.linhas_erro,
        "linhas_dup":     upload.linhas_dup,
        "criado_em":      upload.criado_em.isoformat() if upload.criado_em else None,
        "total_remessas": len(remessas),
        "remessas": [
            {
                "id":       r.id,
                "numero":   r.numero_remessa,
                "cliente":  r.cliente.razao_social if r.cliente else "—",
                "status":   r.status,
            }
            for r in remessas
        ],
    }


# ── HISTÓRICO DE EVENTOS ─────────────────────────────────────────────────────

@app.get("/api/historico", summary="Consulta histórico de eventos")
async def get_historico(
    remessa_id:        Optional[int]  = None,
    transportadora_id: Optional[int]  = None,
    cd_id:             Optional[int]  = None,
    tipo_evento:       Optional[str]  = None,
    gravidade:         Optional[str]  = None,
    apenas_publico:    bool           = False,
    limit:             int            = 100,
    db: AsyncSession = Depends(get_db),
):
    query = select(HistoricoEventos).order_by(HistoricoEventos.timestamp.desc())
    if remessa_id:
        query = query.where(HistoricoEventos.remessa_id == remessa_id)
    if transportadora_id:
        query = query.where(HistoricoEventos.transportadora_id == transportadora_id)
    if cd_id:
        query = query.where(HistoricoEventos.cd_id == cd_id)
    if tipo_evento:
        query = query.where(HistoricoEventos.tipo_evento == tipo_evento)
    if gravidade:
        query = query.where(HistoricoEventos.gravidade == gravidade)
    if apenas_publico:
        # Trava ativa: portal BD não existe ainda — nunca expor eventos internos.
        if settings.VISIBILIDADE_PUBLICA_BLOQUEADA:
            return []
        query = query.where(HistoricoEventos.visibilidade == "publico")
    query = query.limit(min(limit, 500))
    res = await db.execute(query)
    return [serializar_evento(e) for e in res.scalars().all()]


AGENTES_ORIGENS = ["ingestor", "classificador", "montador", "comunicador", "monitor", "resolvedor"]


@app.get("/api/agentes/status", summary="Última execução de cada agente")
async def status_agentes(db: AsyncSession = Depends(get_db)):
    resultado = []
    for origem in AGENTES_ORIGENS:
        res = await db.execute(
            select(HistoricoEventos)
            .where(HistoricoEventos.origem == origem)
            .order_by(HistoricoEventos.timestamp.desc())
            .limit(1)
        )
        evento = res.scalar_one_or_none()
        resultado.append({
            "agente":           origem,
            "ultima_execucao":  evento.timestamp.isoformat() if evento else None,
            "descricao":        evento.descricao if evento else None,
            "resultado":        evento.resultado if evento else None,
            "tipo_evento":      evento.tipo_evento if evento else None,
        })
    return resultado


@app.get("/api/tipos-erro", summary="Lista o catálogo de tipos de erro")
async def listar_tipos_erro(db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(TipoErro).order_by(TipoErro.gravidade.desc(), TipoErro.codigo)
    )
    return [
        {
            "id":            t.id,
            "codigo":        t.codigo,
            "descricao":     t.descricao,
            "gravidade":     t.gravidade,
            "acao_sugerida": t.acao_sugerida,
        }
        for t in res.scalars().all()
    ]


@app.get("/api/erro-acoes", summary="Lista o mapeamento Erro → Ação do Resolvedor")
async def listar_erro_acoes(db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(ErroAcao).where(ErroAcao.ativo == True).order_by(ErroAcao.tipo_erro_codigo)
    )
    return [
        {
            "tipo_erro_codigo":         a.tipo_erro_codigo,
            "acao":                     a.acao,
            "max_tentativas":           a.max_tentativas,
            "intervalo_retry_segundos": a.intervalo_retry_segundos,
        }
        for a in res.scalars().all()
    ]


@app.api_route(
    "/api/resolvedor/executar", methods=["GET", "POST"],
    summary="Varredura periódica do Resolvedor (Vercel Cron, 1x/dia — limite do plano Hobby)",
)
async def executar_resolvedor(request: Request, db: AsyncSession = Depends(get_db)):
    """
    Executado pelo Vercel Cron. Loops infinitos não são viáveis em funções
    serverless (timeout) — a varredura roda uma vez por invocação e o
    agendamento fica a cargo do Cron (vercel.json), não de um loop interno.

    O Vercel Cron invoca este endpoint via GET (com o header
    "Authorization: Bearer <CRON_SECRET>"); POST fica disponível para
    disparo manual/testes com o mesmo segredo.
    """
    auth = request.headers.get("authorization", "")
    if auth != f"Bearer {settings.CRON_SECRET}":
        raise HTTPException(status_code=401, detail="Não autorizado")

    resolvedor = AgenteResolvedor(db)
    return await resolvedor.varrer_erros_pendentes()


# ── ASSISTENTE DE DIAGNÓSTICO ─────────────────────────────────────────────────

class MensagemChatSchema(BaseModel):
    role:    str
    content: str


class AssistenteChatSchema(BaseModel):
    tipo_erro:          Optional[str] = None
    historico_conversa: list[MensagemChatSchema] = []
    mensagem_atual:     str


async def _montar_contexto_sistema(db: AsyncSession) -> dict:
    """Snapshot do estado atual da operação — usado quando o assistente é
    aberto de forma geral (botão flutuante), sem um erro específico em foco."""
    res_status = await db.execute(
        select(Remessa.status, func.count(Remessa.id))
        .where(Remessa.data_extracao == date.today())
        .group_by(Remessa.status)
    )
    remessas_por_status = {status: qtd for status, qtd in res_status.all()}

    res_alertas = await db.execute(
        select(Alerta.severidade, func.count(Alerta.id))
        .where(Alerta.resolvido == False)
        .group_by(Alerta.severidade)
    )
    alertas_por_severidade = {sev: qtd for sev, qtd in res_alertas.all()}

    res_erros = await db.execute(
        select(HistoricoEventos)
        .where(HistoricoEventos.tipo_evento.in_(["erro_sistema", "acao_resolvedor"]))
        .order_by(HistoricoEventos.timestamp.desc())
        .limit(3)
    )
    ultimos_erros = [
        {"timestamp": e.timestamp.strftime("%d/%m/%Y %H:%M"), "descricao": e.descricao}
        for e in res_erros.scalars().all()
    ]

    return {
        "remessas_por_status": remessas_por_status,
        "total_alertas":       sum(alertas_por_severidade.values()),
        "alertas_criticos":    alertas_por_severidade.get("critica", 0),
        "alertas_altos":       alertas_por_severidade.get("alta", 0),
        "ultimos_erros":       ultimos_erros,
    }


@app.post("/api/assistente/chat", summary="Chat contextual do Assistente de Diagnóstico")
async def chat_assistente(
    payload: AssistenteChatSchema,
    db: AsyncSession = Depends(get_db),
):
    tipo_erro_dict = None
    eventos         = None
    contexto_geral  = None

    if payload.tipo_erro:
        res_tipo = await db.execute(select(TipoErro).where(TipoErro.codigo == payload.tipo_erro))
        tipo_erro_obj = res_tipo.scalar_one_or_none()
        if not tipo_erro_obj:
            raise HTTPException(404, f"Tipo de erro não encontrado: {payload.tipo_erro}")

        tipo_erro_dict = {
            "codigo":        tipo_erro_obj.codigo,
            "descricao":     tipo_erro_obj.descricao,
            "gravidade":     tipo_erro_obj.gravidade,
            "acao_sugerida": tipo_erro_obj.acao_sugerida or "Nenhuma ação automática registrada para este erro.",
        }

        # Últimos eventos deste tipo de erro — todas as descrições do Resolvedor
        # seguem o padrão "Erro {codigo} ...", usado aqui para dar contexto de frequência.
        res_hist = await db.execute(
            select(HistoricoEventos)
            .where(HistoricoEventos.descricao.like(f"Erro {payload.tipo_erro} %"))
            .order_by(HistoricoEventos.timestamp.desc())
            .limit(5)
        )
        eventos = [
            {"timestamp": e.timestamp.strftime("%d/%m/%Y %H:%M"), "descricao": e.descricao}
            for e in res_hist.scalars().all()
        ]
    else:
        contexto_geral = await _montar_contexto_sistema(db)

    from agents.agente_assistente import AgenteAssistente
    try:
        agente   = AgenteAssistente()
        resposta = await agente.responder(
            conversa          = [m.model_dump() for m in payload.historico_conversa],
            mensagem          = payload.mensagem_atual,
            tipo_erro         = tipo_erro_dict,
            historico_eventos = eventos,
            contexto_sistema  = contexto_geral,
        )
    except Exception as e:
        logger.error(f"[API] Falha no Assistente de Diagnóstico: {e}")
        raise HTTPException(503, "Assistente de Diagnóstico indisponível no momento. Contate o suporte da Emalog.")

    from core.historico import HistoricoService
    historico = HistoricoService(db)
    await historico.registrar(
        tipo_evento  = "assistente_consultado",
        origem       = "assistente",
        ator_tipo    = "usuario_humano",
        ator_nome    = "operador",
        descricao    = f"Assistente de Diagnóstico consultado sobre {payload.tipo_erro or 'estado geral do sistema'} — pergunta: {payload.mensagem_atual[:200]}",
        resultado    = "sucesso",
        gravidade    = None,
        visibilidade = "interno",
        dados_extra  = {"tipo_erro": payload.tipo_erro},
    )
    await db.commit()

    return {"resposta": resposta}


# ── TRANSPORTADORAS ───────────────────────────────────────────────────────────

@app.get("/api/transportadoras", summary="Lista transportadoras com tabelas de preço")
async def listar_transportadoras(db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(Transportadora)
        .options(selectinload(Transportadora.tabela_precos))
        .where(Transportadora.ativo == True)
        .order_by(Transportadora.nome)
    )
    transportadoras = res.scalars().all()

    return [
        {
            "id":              t.id,
            "codigo":          t.codigo,
            "nome":            t.nome,
            "email_operacoes": t.email_operacoes,
            "integracao":      t.integracao,
            "sla_resposta_h":  t.sla_resposta_h,
            "meta_otif":       t.meta_otif,
            "tabela_precos": [
                {
                    "regiao":          p.regiao,
                    "uf":              p.uf,
                    "classificacao":   p.classificacao,
                    "preco_por_kg":    float(p.preco_por_kg   or 0),
                    "preco_minimo":    float(p.preco_minimo   or 0),
                    "preco_ftl_fixo":  float(p.preco_ftl_fixo or 0),
                    "prazo_frac_dias": p.prazo_frac_dias,
                    "prazo_ftl_dias":  p.prazo_ftl_dias,
                }
                for p in sorted(t.tabela_precos, key=lambda x: x.regiao)
                if p.ativo
            ],
        }
        for t in transportadoras
    ]


class MetaOtifSchema(BaseModel):
    meta_otif: float


@app.patch("/api/transportadoras/{transportadora_id}/meta-otif", summary="Atualiza a meta de OTIF (%) de uma transportadora")
async def atualizar_meta_otif(
    transportadora_id: int,
    payload: MetaOtifSchema,
    db: AsyncSession = Depends(get_db),
):
    if not (0 < payload.meta_otif <= 100):
        raise HTTPException(status_code=422, detail="meta_otif deve estar entre 0 e 100")

    res = await db.execute(select(Transportadora).where(Transportadora.id == transportadora_id))
    transportadora = res.scalar_one_or_none()
    if not transportadora:
        raise HTTPException(status_code=404, detail="Transportadora não encontrada")

    meta_anterior = transportadora.meta_otif
    transportadora.meta_otif = payload.meta_otif

    from core.historico import HistoricoService
    historico = HistoricoService(db)
    await historico.registrar(
        tipo_evento       = "mudanca_status",
        origem            = "usuario",
        ator_tipo         = "usuario_humano",
        ator_nome         = "operador",
        transportadora_id = transportadora.id,
        descricao         = (
            f"Meta de OTIF de {transportadora.nome} alterada de "
            f"{meta_anterior}% para {payload.meta_otif}%"
        ),
        resultado         = "sucesso",
        dados_extra       = {"meta_anterior": meta_anterior, "meta_nova": payload.meta_otif},
    )

    await db.commit()
    return {"id": transportadora.id, "nome": transportadora.nome, "meta_otif": transportadora.meta_otif}


@app.get("/api/transportadoras/estatisticas", summary="Performance das transportadoras via histórico")
async def get_estatisticas_transportadoras(db: AsyncSession = Depends(get_db)):
    cache_key = "transportadoras_estatisticas"
    cached = get_cached(cache_key, ttl_seconds=120)
    if cached is not None:
        return cached

    res_t = await db.execute(
        select(Transportadora).where(Transportadora.ativo == True).order_by(Transportadora.nome)
    )
    transportadoras = res_t.scalars().all()

    # Próximos 3 dias úteis (seg-sex), incluindo hoje se hoje for dia útil —
    # janela usada para "Próximas coletas programadas" de cada transportadora.
    proximos_dias: list[date] = []
    cursor = date.today()
    while len(proximos_dias) < 3:
        if cursor.weekday() < 5:
            proximos_dias.append(cursor)
        cursor += timedelta(days=1)

    transportadora_ids = [t.id for t in transportadoras]

    coletas_por_transp: dict[int, list[dict]] = {}
    programacoes_por_transp: dict[int, int] = {}
    erros_por_transp: dict[int, int] = {}
    progs_por_transp: dict[int, list[ProgramacaoColeta]] = {}
    otif_por_transp: dict[int, dict[str, int]] = {}

    # As 5 agregações abaixo são independentes entre si (era 1 query POR
    # transportadora antes, 5N no total — agora 5 fixas, batched por IN(...)).
    #
    # Nota: cheguei a rodar as 5 em paralelo (sessões concorrentes via
    # asyncio.gather). Revertido — mediu-se contra este Supabase (conexão
    # direta, max_connections=60) resultados inconsistentes (~4-9s, variando
    # run a run) em vez do ganho esperado, provavelmente por contenção do lado
    # do banco. Mantido sequencial; ver mesma nota em _coletar_historico_ondas.
    if transportadora_ids:
        res_coletas = await db.execute(
            select(Onda, PlanoDia.data_plano, CentroDistribuicao.codigo)
            .join(PlanoDia, Onda.plano_id == PlanoDia.id)
            .join(CentroDistribuicao, PlanoDia.cd_id == CentroDistribuicao.id)
            .where(
                Onda.transportadora_id.in_(transportadora_ids),
                Onda.status == "planejada",
                PlanoDia.data_plano.in_(proximos_dias),
            )
            .order_by(PlanoDia.data_plano, Onda.numero_onda)
        )
        for onda, data_plano, cd_codigo in res_coletas.all():
            coletas_por_transp.setdefault(onda.transportadora_id, []).append({
                "onda_id":        onda.id,
                "nome":           onda.nome,
                "cd":             cd_codigo,
                "data":           str(data_plano),
                "horario_coleta": onda.horario_coleta.strftime("%Hh%M") if onda.horario_coleta else None,
                "volume_m3":      float(onda.volume_total_m3 or 0),
            })

        res_prog = await db.execute(
            select(HistoricoEventos.transportadora_id, func.count(HistoricoEventos.id))
            .where(
                and_(
                    HistoricoEventos.transportadora_id.in_(transportadora_ids),
                    HistoricoEventos.tipo_evento       == "decisao_agente",
                    HistoricoEventos.origem            == "comunicador",
                )
            )
            .group_by(HistoricoEventos.transportadora_id)
        )
        programacoes_por_transp = dict(res_prog.all())

        res_erros = await db.execute(
            select(HistoricoEventos.transportadora_id, func.count(HistoricoEventos.id))
            .where(
                and_(
                    HistoricoEventos.transportadora_id.in_(transportadora_ids),
                    HistoricoEventos.tipo_evento       == "erro_sistema",
                )
            )
            .group_by(HistoricoEventos.transportadora_id)
        )
        erros_por_transp = dict(res_erros.all())

        res_progs = await db.execute(
            select(ProgramacaoColeta).where(
                and_(
                    ProgramacaoColeta.transportadora_id.in_(transportadora_ids),
                    ProgramacaoColeta.confirmado_em.isnot(None),
                    ProgramacaoColeta.enviado_em.isnot(None),
                )
            ).order_by(ProgramacaoColeta.id)
        )
        for p in res_progs.scalars().all():
            bucket = progs_por_transp.setdefault(p.transportadora_id, [])
            if len(bucket) < 50:
                bucket.append(p)

        res_otif = await db.execute(
            select(Onda.transportadora_id, Remessa.status, func.count(Remessa.id))
            .join(OndaRemessa, OndaRemessa.remessa_id == Remessa.id)
            .join(Onda, Onda.id == OndaRemessa.onda_id)
            .where(
                Onda.transportadora_id.in_(transportadora_ids),
                Remessa.status.in_(["entregue", "tentativa", "devolvido"]),
            )
            .group_by(Onda.transportadora_id, Remessa.status)
        )
        for transp_id, status, qtd in res_otif.all():
            otif_por_transp.setdefault(transp_id, {})[status] = qtd

    resultado = []
    for t in transportadoras:
        proximas_coletas = coletas_por_transp.get(t.id, [])

        total_programacoes = programacoes_por_transp.get(t.id, 0)
        total_erros        = erros_por_transp.get(t.id, 0)

        total_ops  = total_programacoes + total_erros
        taxa_falha = round(total_erros / total_ops * 100, 1) if total_ops > 0 else 0.0

        progs = progs_por_transp.get(t.id, [])
        deltas = [
            (p.confirmado_em - p.enviado_em).total_seconds() / 3600
            for p in progs
            if p.confirmado_em and p.enviado_em and p.confirmado_em > p.enviado_em
        ]
        tempo_medio_h = round(sum(deltas) / len(deltas), 1) if deltas else None

        contagens_otif = otif_por_transp.get(t.id, {})
        total_fin = sum(contagens_otif.values())
        otif_atual = (
            round(contagens_otif.get("entregue", 0) / total_fin * 100, 1)
            if total_fin > 0 else None
        )
        meta_otif  = t.meta_otif if t.meta_otif is not None else 95.0
        abaixo_meta = otif_atual is not None and otif_atual < meta_otif
        desvio      = round(otif_atual - meta_otif, 1) if otif_atual is not None else None

        resultado.append({
            "id":                      t.id,
            "codigo":                  t.codigo,
            "nome":                    t.nome,
            "sla_resposta_h":          t.sla_resposta_h,
            "total_programacoes":      total_programacoes,
            "total_erros":             total_erros,
            "taxa_falha_pct":          taxa_falha,
            "tempo_medio_resposta_h":  tempo_medio_h,
            "otif_atual":              otif_atual,
            "meta_otif":               meta_otif,
            "abaixo_meta":             abaixo_meta,
            "desvio":                  desvio,
            "proximas_coletas":        proximas_coletas,
        })

    set_cached(cache_key, resultado)
    return resultado


PERIODO_DIAS_RELATORIO_TRANSP = {"semana": 7, "mes": 30, "trimestre": 90}


@app.get("/api/transportadoras/{id}/relatorio", summary="Relatório de performance de uma transportadora")
async def relatorio_transportadora(
    id: int,
    periodo: str = "mes",
    formato: str = "pdf",
    db: AsyncSession = Depends(get_db),
):
    if periodo not in PERIODO_DIAS_RELATORIO_TRANSP:
        raise HTTPException(422, f"periodo inválido — use: {', '.join(PERIODO_DIAS_RELATORIO_TRANSP)}")
    if formato not in ("pdf", "xlsx"):
        raise HTTPException(422, "formato inválido — use 'pdf' ou 'xlsx'")

    res_t = await db.execute(select(Transportadora).where(Transportadora.id == id))
    transp = res_t.scalar_one_or_none()
    if not transp:
        raise HTTPException(404, "Transportadora não encontrada")

    data_fim    = date.today()
    data_inicio = data_fim - timedelta(days=PERIODO_DIAS_RELATORIO_TRANSP[periodo])

    # ── Ondas programadas no período ────────────────────────────────────────
    res_ondas = await db.execute(
        select(Onda, PlanoDia.data_plano, CentroDistribuicao.codigo)
        .join(PlanoDia, Onda.plano_id == PlanoDia.id)
        .join(CentroDistribuicao, PlanoDia.cd_id == CentroDistribuicao.id)
        .where(
            Onda.transportadora_id == transp.id,
            PlanoDia.data_plano >= data_inicio,
            PlanoDia.data_plano <= data_fim,
        )
        .order_by(PlanoDia.data_plano.desc(), Onda.numero_onda)
    )
    ondas_periodo = res_ondas.all()

    contagem_remessas_por_onda: dict[int, int] = {}
    if ondas_periodo:
        res_qtd = await db.execute(
            select(OndaRemessa.onda_id, func.count(OndaRemessa.remessa_id))
            .where(OndaRemessa.onda_id.in_([o.id for o, _, _ in ondas_periodo]))
            .group_by(OndaRemessa.onda_id)
        )
        contagem_remessas_por_onda = dict(res_qtd.all())

    total_ondas     = len(ondas_periodo)
    volume_total_m3 = sum(float(o.volume_total_m3 or 0) for o, _, _ in ondas_periodo)
    peso_total_kg   = sum(float(o.peso_total_kg   or 0) for o, _, _ in ondas_periodo)
    valor_total_nf  = sum(float(o.valor_total_nf  or 0) for o, _, _ in ondas_periodo)
    total_remessas  = sum(contagem_remessas_por_onda.values())

    # ── OTIF do período ──────────────────────────────────────────────────────
    res_otif = await db.execute(
        select(Remessa.status, func.count(Remessa.id))
        .join(OndaRemessa, OndaRemessa.remessa_id == Remessa.id)
        .join(Onda, Onda.id == OndaRemessa.onda_id)
        .join(PlanoDia, Onda.plano_id == PlanoDia.id)
        .where(
            Onda.transportadora_id == transp.id,
            PlanoDia.data_plano >= data_inicio,
            PlanoDia.data_plano <= data_fim,
            Remessa.status.in_(["entregue", "tentativa", "devolvido"]),
        )
        .group_by(Remessa.status)
    )
    contagens_otif = {status: qtd for status, qtd in res_otif.all()}
    total_fin = sum(contagens_otif.values())
    otif_pct  = round(contagens_otif.get("entregue", 0) / total_fin * 100, 1) if total_fin > 0 else None
    meta_otif = transp.meta_otif if transp.meta_otif is not None else 95.0

    # ── Erros do período ──────────────────────────────────────────────────────
    res_erros = await db.execute(
        select(HistoricoEventos)
        .where(
            HistoricoEventos.transportadora_id == transp.id,
            HistoricoEventos.tipo_evento        == "erro_sistema",
            HistoricoEventos.timestamp >= datetime.combine(data_inicio, dtime.min),
            HistoricoEventos.timestamp <= datetime.combine(data_fim, dtime.max),
        )
        .order_by(HistoricoEventos.timestamp.desc())
        .limit(50)
    )
    erros = res_erros.scalars().all()

    nome_arquivo = f"relatorio_{transp.codigo}_{periodo}_{data_fim:%Y%m%d}"

    if formato == "xlsx":
        return _relatorio_transportadora_xlsx(
            transp, periodo, data_inicio, data_fim, total_ondas, total_remessas,
            volume_total_m3, peso_total_kg, valor_total_nf, otif_pct, meta_otif,
            ondas_periodo, contagem_remessas_por_onda, erros, nome_arquivo,
        )
    return _relatorio_transportadora_pdf(
        transp, periodo, data_inicio, data_fim, total_ondas, total_remessas,
        volume_total_m3, peso_total_kg, valor_total_nf, otif_pct, meta_otif,
        ondas_periodo, contagem_remessas_por_onda, erros, nome_arquivo,
    )


def _relatorio_transportadora_pdf(
    transp, periodo, data_inicio, data_fim, total_ondas, total_remessas,
    volume_total_m3, peso_total_kg, valor_total_nf, otif_pct, meta_otif,
    ondas_periodo, contagem_remessas_por_onda, erros, nome_arquivo,
):
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
    )

    BD_BLUE   = colors.HexColor("#1A3F6F")
    GRAY_ROW  = colors.HexColor("#F5F7FA")
    GRAY_LINE = colors.HexColor("#E5E7EB")
    GREEN_BG  = colors.HexColor("#D1FAE5")
    RED_BG    = colors.HexColor("#FEE2E2")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.8*cm, bottomMargin=1.8*cm,
    )
    styles = getSampleStyleSheet()

    def st(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    title_st = st("T", textColor=BD_BLUE, fontSize=16, fontName="Helvetica-Bold", spaceAfter=2)
    sub_st   = st("S", textColor=colors.HexColor("#6B7280"), fontSize=9, spaceAfter=10)
    h3_st    = st("H3", textColor=BD_BLUE, fontSize=10, fontName="Helvetica-Bold", spaceBefore=10, spaceAfter=4)
    cell_st  = st("C", fontSize=7, leading=9)
    foot_st  = st("F", fontSize=6, textColor=colors.HexColor("#9CA3AF"), alignment=TA_CENTER)

    def tbl_style():
        return TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), BD_BLUE),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 7),
            ("FONTSIZE",      (0, 1), (-1, -1), 7),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, GRAY_ROW]),
            ("GRID",          (0, 0), (-1, -1), 0.4, GRAY_LINE),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])

    story = [
        Paragraph(f"Relatório de Performance — {transp.nome}", title_st),
        Paragraph(
            f"Período: {periodo} · {data_inicio:%d/%m/%Y} a {data_fim:%d/%m/%Y} · "
            f"Gerado em {datetime.now():%d/%m/%Y %Hh%M}",
            sub_st,
        ),
        HRFlowable(width="100%", thickness=0.5, color=GRAY_LINE),
    ]

    # ── Resumo ────────────────────────────────────────────────────────────
    otif_txt = f"{otif_pct:.1f}%" if otif_pct is not None else "sem dados"
    abaixo   = otif_pct is not None and otif_pct < meta_otif
    resumo_rows = [
        ["Ondas programadas", str(total_ondas), "Remessas transportadas", str(total_remessas)],
        ["Volume total (m³)", f"{volume_total_m3:.2f}", "Peso total (kg)", f"{peso_total_kg:.1f}"],
        ["Valor total NF (R$)", f"{valor_total_nf:,.2f}", "OTIF vs. meta", f"{otif_txt} / {meta_otif:.1f}%"],
    ]
    t_resumo = Table(resumo_rows, colWidths=[4.2*cm, 3*cm, 4.2*cm, 3*cm])
    t_resumo.setStyle(TableStyle([
        ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",   (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("TEXTCOLOR",  (0, 0), (0, -1), colors.HexColor("#4B5563")),
        ("TEXTCOLOR",  (2, 0), (2, -1), colors.HexColor("#4B5563")),
        ("BACKGROUND", (3, 2), (3, 2), RED_BG if abaixo else GREEN_BG),
        ("GRID",       (0, 0), (-1, -1), 0.4, GRAY_LINE),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(t_resumo)

    # ── Ondas programadas no período ─────────────────────────────────────────
    story.append(Paragraph("Ondas programadas no período", h3_st))
    if ondas_periodo:
        header = ["Data", "CD", "Onda", "Tipo", "Status", "Remessas", "Vol. m³"]
        rows = [header]
        for onda, data_plano, cd_codigo in ondas_periodo:
            rows.append([
                data_plano.strftime("%d/%m/%Y"),
                cd_codigo,
                onda.nome,
                (onda.tipo or "—").upper(),
                onda.status,
                str(contagem_remessas_por_onda.get(onda.id, 0)),
                f"{float(onda.volume_total_m3 or 0):.1f}",
            ])
        t_ondas = Table(rows, colWidths=[2.2*cm, 1.3*cm, 5.5*cm, 1.8*cm, 2*cm, 2*cm, 2*cm], repeatRows=1)
        t_ondas.setStyle(tbl_style())
        story.append(t_ondas)
    else:
        story.append(Paragraph("Nenhuma onda programada no período.", cell_st))

    # ── Erros no período ──────────────────────────────────────────────────────
    story.append(Paragraph("Erros registrados no período", h3_st))
    if erros:
        header = ["Data", "Descrição", "Gravidade"]
        rows = [header]
        for e in erros:
            rows.append([
                e.timestamp.strftime("%d/%m/%Y %Hh%M"),
                Paragraph(e.descricao, cell_st),
                e.gravidade or "—",
            ])
        t_erros = Table(rows, colWidths=[3*cm, 11.8*cm, 2*cm], repeatRows=1)
        t_erros.setStyle(tbl_style())
        story.append(t_erros)
    else:
        story.append(Paragraph("Nenhum erro registrado no período.", cell_st))

    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.4, color=GRAY_LINE))
    story.append(Paragraph("Becton Dickinson — PEO-BD v1.0  |  Documento gerado automaticamente", foot_st))

    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={nome_arquivo}.pdf"},
    )


def _relatorio_transportadora_xlsx(
    transp, periodo, data_inicio, data_fim, total_ondas, total_remessas,
    volume_total_m3, peso_total_kg, valor_total_nf, otif_pct, meta_otif,
    ondas_periodo, contagem_remessas_por_onda, erros, nome_arquivo,
):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"

    azul_bd    = PatternFill("solid", fgColor="1A3F6F")
    header_font = Font(bold=True, color="FFFFFF")
    bold        = Font(bold=True)

    ws["A1"] = f"Relatório de Performance — {transp.nome}"
    ws["A1"].font = Font(bold=True, size=14, color="1A3F6F")
    ws["A2"] = f"Período: {periodo} · {data_inicio:%d/%m/%Y} a {data_fim:%d/%m/%Y}"
    ws["A2"].font = Font(italic=True, color="6B7280")
    ws["A3"] = f"Gerado em {datetime.now():%d/%m/%Y %Hh%M}"
    ws["A3"].font = Font(italic=True, size=8, color="9CA3AF")

    resumo = [
        ("Ondas programadas",     total_ondas),
        ("Remessas transportadas", total_remessas),
        ("Volume total (m³)",     round(volume_total_m3, 2)),
        ("Peso total (kg)",       round(peso_total_kg, 1)),
        ("Valor total NF (R$)",   round(valor_total_nf, 2)),
        ("OTIF do período (%)",   otif_pct if otif_pct is not None else "sem dados"),
        ("Meta OTIF (%)",         meta_otif),
    ]
    linha = 5
    for label, valor in resumo:
        ws.cell(row=linha, column=1, value=label).font = bold
        ws.cell(row=linha, column=2, value=valor)
        linha += 1

    linha += 1
    ws.cell(row=linha, column=1, value="Ondas programadas no período").font = Font(bold=True, size=12, color="1A3F6F")
    linha += 1
    header_ondas = ["Data", "CD", "Onda", "Tipo", "Status", "Remessas", "Vol. m³"]
    for col, texto in enumerate(header_ondas, start=1):
        c = ws.cell(row=linha, column=col, value=texto)
        c.font = header_font
        c.fill = azul_bd
    linha += 1
    for onda, data_plano, cd_codigo in ondas_periodo:
        ws.append([
            data_plano.strftime("%d/%m/%Y"), cd_codigo, onda.nome,
            (onda.tipo or "—").upper(), onda.status,
            contagem_remessas_por_onda.get(onda.id, 0),
            float(onda.volume_total_m3 or 0),
        ])
        linha += 1

    linha += 1
    ws.cell(row=linha, column=1, value="Erros registrados no período").font = Font(bold=True, size=12, color="1A3F6F")
    linha += 1
    header_erros = ["Data", "Descrição", "Gravidade"]
    for col, texto in enumerate(header_erros, start=1):
        c = ws.cell(row=linha, column=col, value=texto)
        c.font = header_font
        c.fill = azul_bd
    linha += 1
    for e in erros:
        ws.append([e.timestamp.strftime("%d/%m/%Y %Hh%M"), e.descricao, e.gravidade or "—"])

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome_arquivo}.xlsx"},
    )


@app.get("/api/transportadoras/cotacao", summary="Cotação consolidada por rota/região")
async def cotacao_transportadoras(
    cd_codigo: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    cd_id = None
    if cd_codigo:
        r = await db.execute(select(CentroDistribuicao).where(CentroDistribuicao.codigo == cd_codigo))
        cd = r.scalar_one_or_none()
        if cd:
            cd_id = cd.id

    # Agrupa remessas em aberto por região
    filtros = [Remessa.status == "novo"]
    if cd_id:
        filtros.append(Remessa.cd_id == cd_id)

    res_regioes = await db.execute(
        select(
            Cliente.regiao,
            func.count(Remessa.id).label("total_remessas"),
            func.sum(Remessa.peso_kg).label("peso_total"),
            func.sum(Remessa.volume_m3).label("volume_total"),
        )
        .join(Cliente, Remessa.cliente_id == Cliente.id)
        .where(and_(*filtros))
        .group_by(Cliente.regiao)
        .order_by(Cliente.regiao)
    )
    regioes = res_regioes.all()

    # Busca todas as transportadoras com preços
    res_transp = await db.execute(
        select(Transportadora)
        .options(selectinload(Transportadora.tabela_precos))
        .where(Transportadora.ativo == True)
    )
    transportadoras = res_transp.scalars().all()

    resultado = []
    for reg in regioes:
        regiao      = reg.regiao or "nao_mapeada"
        peso_total  = float(reg.peso_total  or 0)
        vol_total   = float(reg.volume_total or 0)
        ftl_viavel  = vol_total >= settings.LIMIAR_FTL_M3

        cotacoes = []
        for t in transportadoras:
            preco_frac = None
            preco_ftl  = None
            prazo      = None

            for p in t.tabela_precos:
                if not p.ativo:
                    continue
                if p.regiao != regiao:
                    continue
                if p.tipo_servico == "fracionado":
                    custo = peso_total * float(p.preco_por_kg or 0)
                    preco_frac = max(custo, float(p.preco_minimo or 0))
                    prazo = p.prazo_dias
                elif p.tipo_servico == "ftl":
                    preco_ftl = float(p.preco_ftl_fixo or 0)
                    if prazo is None:
                        prazo = p.prazo_dias

            if preco_frac is None and preco_ftl is None:
                continue

            if ftl_viavel and preco_ftl is not None:
                melhor_opcao  = "ftl" if preco_ftl < (preco_frac or 9e9) else "fracionado"
                custo_melhor  = min(p for p in [preco_frac, preco_ftl] if p is not None)
            else:
                melhor_opcao  = "fracionado"
                custo_melhor  = preco_frac or 0

            cotacoes.append({
                "transportadora":  t.nome,
                "codigo":          t.codigo,
                "custo_fracionado": round(preco_frac or 0, 2),
                "custo_ftl":        round(preco_ftl  or 0, 2) if preco_ftl else None,
                "melhor_opcao":    melhor_opcao,
                "custo_melhor":    round(custo_melhor, 2),
                "prazo_dias":      prazo,
            })

        cotacoes.sort(key=lambda x: x["custo_melhor"])

        melhor = cotacoes[0]["codigo"] if cotacoes else None
        economia = round(
            (cotacoes[-1]["custo_melhor"] - cotacoes[0]["custo_melhor"]) if len(cotacoes) > 1 else 0, 2
        )

        resultado.append({
            "regiao":              regiao,
            "total_remessas":      reg.total_remessas,
            "peso_total_kg":       round(peso_total, 1),
            "volume_total_m3":     round(vol_total, 2),
            "ftl_viavel":          ftl_viavel,
            "cotacoes":            cotacoes,
            "melhor_transportadora": melhor,
            "economia_vs_pior":    economia,
        })

    return resultado


# ── COTAÇÃO DE TRANSPORTADORAS ───────────────────────────────────────────────

@app.get("/api/cotacao/template", summary="Download do formulário em branco para transportadoras")
async def download_template():
    """Retorna o formulário .xlsx para ser enviado às transportadoras parceiras."""
    from scripts.gerar_planilha_cotacao import gerar_formulario
    import tempfile

    tmp = Path(tempfile.mkdtemp())
    caminho = gerar_formulario(tmp)
    return FileResponse(
        path     = caminho,
        filename = caminho.name,
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/api/cotacao/importar", summary="Importa planilha preenchida pela transportadora")
async def importar_cotacao(
    arquivo:  UploadFile = File(...),
    usuario:  str        = Form(...),
    db:       AsyncSession = Depends(get_db),
):
    """
    Lê o formulário .xlsx preenchido pela transportadora e grava/atualiza
    a tabela tabela_preco_transportadoras no banco.
    Retorna resumo: transportadora identificada, linhas importadas, erros.
    """
    import openpyxl
    from datetime import date as date_type

    # Posições fixas definidas em gerar_planilha_cotacao.py
    DATA_START   = 15   # linha 14 = header, 15+ = dados
    CEL_NOME_ROW, CEL_NOME_COL = 3, 3   # C3 = nome transportadora
    CEL_VAL_FIM_ROW, CEL_VAL_FIM_COL = 4, 9   # I4 = validade até

    COL_MACRO     = 2   # B
    COL_ESTADO    = 3   # C
    COL_UF        = 4   # D
    COL_CLASS     = 5   # E
    COL_COBRE     = 6   # F
    COL_PRECO_KG  = 7   # G
    COL_PESO_MIN  = 8   # H
    COL_PRECO_MIN = 9   # I
    COL_PRAZO_FRAC= 10  # J
    COL_PRECO_FTL = 11  # K
    COL_PRAZO_FTL = 12  # L
    COL_ADVALOREM = 13  # M
    COL_GRIS      = 14  # N
    COL_SLA       = 15  # O
    COL_OBS       = 16  # P

    def _val(ws, row, col):
        v = ws.cell(row=row, column=col).value
        return v if v not in (None, "N/A", "N/a", "") else None

    def _float(v):
        try:
            return float(str(v).replace(",", "."))
        except Exception:
            return None

    def _int(v):
        try:
            return int(float(str(v).replace(",", ".")))
        except Exception:
            return None

    # Salva temporariamente
    tmp_path = settings.UPLOAD_DIR / f"cotacao_{arquivo.filename}"
    with tmp_path.open("wb") as f:
        import shutil
        shutil.copyfileobj(arquivo.file, f)

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Arquivo inválido: {e}")

    ws = wb.active

    # Lê identificação
    nome_transp = _val(ws, CEL_NOME_ROW, CEL_NOME_COL)
    if not nome_transp:
        raise HTTPException(422, "Campo 'Transportadora' não preenchido na planilha (célula C3).")

    val_ini_raw = None
    val_fim_raw = _val(ws, CEL_VAL_FIM_ROW, CEL_VAL_FIM_COL)

    def _parse_date(v):
        if isinstance(v, (date_type, datetime)):
            return v if isinstance(v, date_type) else v.date()
        if v:
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    return datetime.strptime(str(v).strip(), fmt).date()
                except ValueError:
                    pass
        return None

    validade_ini = _parse_date(val_ini_raw)
    validade_fim = _parse_date(val_fim_raw)

    # Localiza ou cria Transportadora
    res_t = await db.execute(
        select(Transportadora).where(Transportadora.nome.ilike(f"%{str(nome_transp).strip()}%"))
    )
    transp = res_t.scalar_one_or_none()

    if not transp:
        # Cria nova transportadora sem CD vinculado (pode ser associada depois)
        transp = Transportadora(
            codigo          = str(nome_transp).upper()[:20].replace(" ", "_"),
            nome            = str(nome_transp).strip(),
            email_operacoes = None,
            integracao      = "email",
            sla_resposta_h  = 2,
            ativo           = True,
        )
        db.add(transp)
        await db.flush()

    # Lê linhas de dados
    importadas = erros = ignoradas = 0
    erros_log = []

    # Acumula macro/estado/uf para linhas onde a célula está vazia (merge visual)
    macro_atual  = ""
    estado_atual = ""
    uf_atual     = ""

    row = DATA_START
    while True:
        macro_raw = _val(ws, row, COL_MACRO)
        uf_raw    = _val(ws, row, COL_UF)

        # Célula vazia = continuação da macro/estado anterior
        if macro_raw:
            macro_atual = str(macro_raw).strip()
        if uf_raw:
            uf_atual    = str(uf_raw).strip()

        estado_raw = _val(ws, row, COL_ESTADO)
        if estado_raw:
            estado_atual = str(estado_raw).strip()

        classif = _val(ws, row, COL_CLASS)
        cobertura = _val(ws, row, COL_COBRE)

        # Linha vazia ou sem UF depois de muito espaço = fim da tabela
        if not uf_atual and not classif:
            if row > DATA_START + 5:
                break
            row += 1
            continue

        if not classif:
            row += 1
            continue

        cobertura_bool = str(cobertura or "").strip().lower() not in ("nao", "não", "no", "n", "false", "0")

        try:
            # Upsert por (transportadora_id, uf, classificacao)
            res_p = await db.execute(
                select(TabelaPrecoTransportadora).where(
                    TabelaPrecoTransportadora.transportadora_id == transp.id,
                    TabelaPrecoTransportadora.uf                == uf_atual,
                    TabelaPrecoTransportadora.classificacao     == str(classif).strip(),
                )
            )
            preco = res_p.scalar_one_or_none()

            if not preco:
                preco = TabelaPrecoTransportadora(
                    transportadora_id = transp.id,
                    macro_regiao      = macro_atual,
                    estado            = estado_atual,
                    uf                = uf_atual,
                    classificacao     = str(classif).strip(),
                )
                db.add(preco)
            else:
                preco.macro_regiao = macro_atual
                preco.estado       = estado_atual

            preco.cobertura         = cobertura_bool
            preco.preco_por_kg      = _float(_val(ws, row, COL_PRECO_KG))
            preco.peso_minimo_kg    = _float(_val(ws, row, COL_PESO_MIN))
            preco.preco_minimo      = _float(_val(ws, row, COL_PRECO_MIN))
            preco.prazo_frac_dias   = _int(_val(ws, row, COL_PRAZO_FRAC))
            preco.preco_ftl_fixo    = _float(_val(ws, row, COL_PRECO_FTL))
            preco.prazo_ftl_dias    = _int(_val(ws, row, COL_PRAZO_FTL))
            preco.ad_valorem_pct    = _float(_val(ws, row, COL_ADVALOREM))
            preco.gris_pct          = _float(_val(ws, row, COL_GRIS))
            preco.sla_confirmacao_h = _int(_val(ws, row, COL_SLA))
            preco.observacoes       = _val(ws, row, COL_OBS)
            preco.validade_inicio   = validade_ini
            preco.validade_fim      = validade_fim
            preco.ativo             = True
            preco.atualizado_em     = datetime.utcnow()

            importadas += 1
        except Exception as e:
            erros += 1
            erros_log.append(f"Linha {row} ({uf_atual}/{classif}): {e}")

        row += 1

    await db.commit()
    logger.info(
        f"[Cotação] Importada: {nome_transp} | {importadas} rotas | {erros} erros | user={usuario}"
    )

    return {
        "transportadora":    str(nome_transp),
        "transportadora_id": transp.id,
        "validade_inicio":   str(validade_ini) if validade_ini else None,
        "validade_fim":      str(validade_fim) if validade_fim else None,
        "linhas_importadas": importadas,
        "erros":             erros,
        "log_erros":         erros_log[:20],
    }


@app.get("/api/cotacao/comparativo", summary="Comparativo de transportadoras por rota (uso interno BD)")
async def comparativo_cotacao(
    regiao:       Optional[str] = None,
    tipo_servico: Optional[str] = None,
    peso_kg:      Optional[float] = None,
    valor_nf:     Optional[float] = None,
    db: AsyncSession = Depends(get_db),
):
    """
    Retorna comparativo de custo por transportadora para uma rota/serviço.
    Se peso_kg e valor_nf forem informados, calcula custo estimado real.
    """
    filtros = [
        TabelaPrecoTransportadora.ativo     == True,
        TabelaPrecoTransportadora.cobertura == True,
    ]
    if regiao:
        filtros.append(TabelaPrecoTransportadora.regiao == regiao)
    if tipo_servico:
        filtros.append(TabelaPrecoTransportadora.tipo_servico == tipo_servico)

    res = await db.execute(
        select(TabelaPrecoTransportadora)
        .options(selectinload(TabelaPrecoTransportadora.transportadora))
        .where(*filtros)
        .order_by(TabelaPrecoTransportadora.regiao, TabelaPrecoTransportadora.tipo_servico)
    )
    precos = res.scalars().all()

    def _calcular_custo(p: TabelaPrecoTransportadora, kg: float, nf: float) -> float | None:
        if p.tipo_servico == "ftl":
            return float(p.preco_ftl_fixo or 0) or None
        custo_kg   = float(p.preco_por_kg   or 0) * kg
        custo_adv  = float(p.ad_valorem_pct or 0) * nf
        custo_gris = float(p.gris_pct        or 0) * nf
        custo      = custo_kg + custo_adv + custo_gris
        return round(max(custo, float(p.preco_minimo or 0)), 2)

    # Agrupa por rota
    rotas: dict[str, dict] = {}
    for p in precos:
        chave = f"{p.regiao}||{p.tipo_servico}"
        if chave not in rotas:
            rotas[chave] = {
                "regiao":       p.regiao,
                "tipo_servico": p.tipo_servico,
                "cotacoes":     [],
            }

        custo_estimado = None
        if peso_kg is not None and valor_nf is not None:
            custo_estimado = _calcular_custo(p, peso_kg, valor_nf)

        rotas[chave]["cotacoes"].append({
            "transportadora_id":   p.transportadora_id,
            "transportadora":      p.transportadora.nome if p.transportadora else "—",
            "preco_por_kg":        float(p.preco_por_kg   or 0),
            "preco_minimo":        float(p.preco_minimo   or 0),
            "preco_ftl_fixo":      float(p.preco_ftl_fixo or 0),
            "ad_valorem_pct":      float(p.ad_valorem_pct or 0),
            "gris_pct":            float(p.gris_pct        or 0),
            "prazo_dias":          p.prazo_dias,
            "sla_confirmacao_h":   p.sla_confirmacao_h,
            "custo_estimado":      custo_estimado,
            "validade_fim":        str(p.validade_fim) if p.validade_fim else None,
        })

    resultado = []
    for dados in rotas.values():
        cotacoes = dados["cotacoes"]
        cotacoes_com_custo = [c for c in cotacoes if c["custo_estimado"] is not None]

        if cotacoes_com_custo:
            cotacoes_com_custo.sort(key=lambda x: (x["custo_estimado"], x["prazo_dias"] or 999))
            melhor = cotacoes_com_custo[0]
            economia = round(
                cotacoes_com_custo[-1]["custo_estimado"] - cotacoes_com_custo[0]["custo_estimado"], 2
            ) if len(cotacoes_com_custo) > 1 else 0
        else:
            cotacoes.sort(key=lambda x: (x["preco_por_kg"] or x["preco_ftl_fixo"] or 0))
            melhor = cotacoes[0] if cotacoes else None
            economia = 0

        resultado.append({
            "regiao":                dados["regiao"],
            "tipo_servico":          dados["tipo_servico"],
            "total_transportadoras": len(cotacoes),
            "melhor_transportadora": melhor["transportadora"] if melhor else None,
            "melhor_custo_estimado": melhor["custo_estimado"] if melhor else None,
            "economia_vs_pior":      economia,
            "cotacoes":              cotacoes,
        })

    resultado.sort(key=lambda x: (x["regiao"], x["tipo_servico"]))
    return resultado


# ── RELATÓRIO PDF — ONDAS DO DIA ─────────────────────────────────────────────

@app.get("/api/relatorio/pdf", summary="Exportar PDF das Ondas do Dia")
async def relatorio_pdf(
    cd_codigo: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    import io
    from datetime import datetime
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable, KeepTogether,
    )

    BD_BLUE    = colors.HexColor("#1A3F6F")
    BD_LIGHT   = colors.HexColor("#E8F0F8")
    GRAY_ROW   = colors.HexColor("#F5F7FA")
    GRAY_LINE  = colors.HexColor("#E5E7EB")
    GREEN      = colors.HexColor("#065F46")
    GREEN_BG   = colors.HexColor("#D1FAE5")
    ORANGE_BG  = colors.HexColor("#FEF3C7")
    RED_BG     = colors.HexColor("#FEE2E2")

    hoje = date.today()

    # ── Busca planos e ondas do dia ───────────────────────────────────────────
    filtro_plano = [PlanoDia.data_plano == hoje]
    if cd_codigo:
        res_cd = await db.execute(
            select(CentroDistribuicao).where(CentroDistribuicao.codigo == cd_codigo)
        )
        cd_obj = res_cd.scalar_one_or_none()
        if cd_obj:
            filtro_plano.append(PlanoDia.cd_id == cd_obj.id)

    res_planos = await db.execute(
        select(PlanoDia)
        .options(selectinload(PlanoDia.cd))
        .where(and_(*filtro_plano))
        .order_by(PlanoDia.criado_em.desc())
    )
    planos = res_planos.scalars().all()

    # Para cada onda, busca remessas
    ondas_com_remessas = []
    for plano in planos:
        res_ondas = await db.execute(
            select(Onda)
            .options(selectinload(Onda.transportadora), selectinload(Onda.veiculo))
            .where(Onda.plano_id == plano.id)
            .order_by(Onda.numero_onda)
        )
        for onda in res_ondas.scalars().all():
            res_rem = await db.execute(
                select(Remessa)
                .options(selectinload(Remessa.cliente))
                .join(OndaRemessa, OndaRemessa.remessa_id == Remessa.id)
                .where(OndaRemessa.onda_id == onda.id)
                .order_by(Remessa.prioridade, Remessa.numero_remessa)
            )
            remessas = res_rem.scalars().all()
            ondas_com_remessas.append({
                "onda":     onda,
                "plano":    plano,
                "remessas": remessas,
            })

    # ── Monta PDF ─────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.8*cm, bottomMargin=1.8*cm,
    )
    styles = getSampleStyleSheet()

    def st(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    title_st = st("T", textColor=BD_BLUE, fontSize=16, fontName="Helvetica-Bold", spaceAfter=2)
    sub_st   = st("S", textColor=colors.HexColor("#6B7280"), fontSize=8, spaceAfter=0)
    h3_st    = st("H3", textColor=BD_BLUE, fontSize=10, fontName="Helvetica-Bold",
                  spaceBefore=8, spaceAfter=3)
    cell_st  = st("C", fontSize=7, leading=9)
    foot_st  = st("F", fontSize=6, textColor=colors.HexColor("#9CA3AF"), alignment=TA_CENTER)
    just_st  = st("J", fontSize=7, leading=10, textColor=colors.HexColor("#4B5563"),
                  spaceAfter=4, fontName="Helvetica-Oblique")

    def tbl_style(rows, header_cols=None):
        ts = TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), BD_BLUE),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 7),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, GRAY_ROW]),
            ("GRID",          (0, 0), (-1, -1), 0.3, GRAY_LINE),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE",      (0, 1), (-1, -1), 7),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
        ])
        return ts

    story = []

    # ── Cabeçalho do documento ────────────────────────────────────────────────
    cd_label = f"CD: {cd_codigo}" if cd_codigo else "Todos os CDs"
    story.append(Paragraph("PEO-BD — Programação de Ondas do Dia", title_st))
    story.append(Paragraph(
        f"{cd_label}  ·  {hoje.strftime('%d/%m/%Y')}  ·  Gerado às {datetime.now().strftime('%H:%M')}",
        sub_st,
    ))
    story.append(Spacer(1, 0.2*cm))
    story.append(HRFlowable(width="100%", thickness=2, color=BD_BLUE))
    story.append(Spacer(1, 0.3*cm))

    # ── Resumo do dia ─────────────────────────────────────────────────────────
    total_ondas    = len(ondas_com_remessas)
    total_remessas = sum(len(o["remessas"]) for o in ondas_com_remessas)
    total_vol      = sum(float(o["onda"].volume_total_m3 or 0) for o in ondas_com_remessas)
    total_valor    = sum(float(o["onda"].valor_total_nf  or 0) for o in ondas_com_remessas)
    total_peso     = sum(float(o["onda"].peso_total_kg   or 0) for o in ondas_com_remessas)

    resumo_rows = [[
        "Ondas planejadas", "Total de remessas", "Volume total (m³)", "Peso total (kg)", "Valor total NF",
    ], [
        str(total_ondas),
        str(total_remessas),
        f"{total_vol:.2f}",
        f"{total_peso:.0f}",
        f"R$ {total_valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
    ]]
    t_res = Table(resumo_rows, colWidths=[3.5*cm]*5)
    ts_res = TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), BD_LIGHT),
        ("TEXTCOLOR",     (0, 0), (-1, 0), BD_BLUE),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 7),
        ("FONTNAME",      (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 1), (-1, 1), 11),
        ("TEXTCOLOR",     (0, 1), (-1, 1), BD_BLUE),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",          (0, 0), (-1, -1), 0.3, GRAY_LINE),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("ROUNDEDCORNERS", [3]),
    ])
    t_res.setStyle(ts_res)
    story.append(t_res)
    story.append(Spacer(1, 0.4*cm))

    # ── Tabela resumo de ondas ────────────────────────────────────────────────
    if ondas_com_remessas:
        story.append(Paragraph("Resumo das Ondas", h3_st))
        onda_hdr = ["Onda", "CD", "Transportadora", "Tipo", "Coleta", "Remessas",
                    "Vol m³", "Peso kg", "Ocup.%", "Valor NF", "Status"]
        onda_rows = [onda_hdr]
        for item in ondas_com_remessas:
            o = item["onda"]
            onda_rows.append([
                o.nome or f"Onda {o.numero_onda}",
                item["plano"].cd.codigo if item["plano"].cd else "—",
                o.transportadora.nome if o.transportadora else "—",
                o.tipo or "—",
                o.horario_coleta.strftime("%Hh%M") if o.horario_coleta else "—",
                str(len(item["remessas"])),
                f"{float(o.volume_total_m3 or 0):.2f}",
                f"{float(o.peso_total_kg  or 0):.0f}",
                f"{float(o.ocupacao_pct   or 0):.0f}%",
                f"R$ {float(o.valor_total_nf or 0):,.0f}".replace(",", "."),
                o.status or "—",
            ])

        col_w = [2.8*cm, 1.2*cm, 3*cm, 1.8*cm, 1.5*cm, 1.8*cm,
                 1.6*cm, 1.6*cm, 1.5*cm, 2.2*cm, 1.5*cm]
        t_ondas = Table(onda_rows, colWidths=col_w)
        ts = tbl_style(onda_rows)
        # Colorir status
        for i, row in enumerate(onda_rows[1:], start=1):
            status = row[-1].lower()
            if status == "executado":
                ts.add("BACKGROUND", (-1, i), (-1, i), GREEN_BG)
                ts.add("TEXTCOLOR",  (-1, i), (-1, i), GREEN)
            elif status in ("atrasado", "pendente"):
                ts.add("BACKGROUND", (-1, i), (-1, i), ORANGE_BG)
        t_ondas.setStyle(ts)
        story.append(KeepTogether(t_ondas))
        story.append(Spacer(1, 0.5*cm))

    # ── Detalhe por onda ─────────────────────────────────────────────────────
    for item in ondas_com_remessas:
        o        = item["onda"]
        remessas = item["remessas"]
        if not remessas:
            continue

        nome_onda = o.nome or f"Onda {o.numero_onda}"
        transp    = o.transportadora.nome if o.transportadora else "—"
        coleta    = o.horario_coleta.strftime("%Hh%M") if o.horario_coleta else "—"
        ocup      = f"{float(o.ocupacao_pct or 0):.0f}%"

        story.append(Paragraph(
            f"{nome_onda}  ·  {transp}  ·  Coleta: {coleta}  ·  Ocupação: {ocup}",
            h3_st,
        ))
        if o.justificativa:
            story.append(Paragraph(
                "Por que essas decisões: " + o.justificativa.replace("\n", "<br/>"),
                just_st,
            ))

        rem_hdr = ["Remessa", "Cliente", "Cidade / UF", "Tipo", "Vol m³", "Peso kg", "Valor NF", "Prioridade", "Status"]
        rem_rows = [rem_hdr]
        for r in remessas:
            cliente_nome = (r.cliente.razao_social[:30] if r.cliente else "—")
            cidade_uf    = ""
            if r.cliente:
                cidade_uf = f"{r.cliente.cidade or ''}/{r.cliente.uf or ''}".strip("/")
            rem_rows.append([
                r.numero_remessa or "—",
                cliente_nome,
                cidade_uf or "—",
                r.cliente.tipo[:8] if r.cliente and r.cliente.tipo else "—",
                f"{float(r.volume_m3 or 0):.2f}",
                f"{float(r.peso_kg   or 0):.0f}",
                f"R$ {float(r.valor_nf or 0):,.0f}".replace(",", "."),
                (r.prioridade or "normal").upper(),
                (r.status or "—").replace("_", " "),
            ])

        col_w2 = [2.2*cm, 3.8*cm, 2.5*cm, 1.5*cm, 1.4*cm, 1.4*cm, 2*cm, 1.8*cm, 1.9*cm]
        t_rem = Table(rem_rows, colWidths=col_w2)
        ts2 = tbl_style(rem_rows)
        # Destacar prioridade crítica
        for i, row in enumerate(rem_rows[1:], start=1):
            pri = row[-2]
            if pri == "CRITICA":
                ts2.add("BACKGROUND", (0, i), (-1, i), RED_BG)
            elif pri == "ALTA":
                ts2.add("BACKGROUND", (0, i), (-1, i), ORANGE_BG)
        t_rem.setStyle(ts2)
        story.append(KeepTogether(t_rem))
        story.append(Spacer(1, 0.5*cm))

    # ── Sem ondas ─────────────────────────────────────────────────────────────
    if not ondas_com_remessas:
        story.append(Paragraph(
            f"Nenhuma onda planejada para {hoje.strftime('%d/%m/%Y')}.",
            st("NA", textColor=colors.HexColor("#6B7280"), fontSize=10, alignment=TA_CENTER),
        ))

    # ── Rodapé ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.6*cm))
    story.append(HRFlowable(width="100%", thickness=0.4, color=GRAY_LINE))
    story.append(Paragraph("Becton Dickinson — PEO-BD v1.0  |  Documento gerado automaticamente", foot_st))

    doc.build(story)
    buffer.seek(0)

    filename = f"ondas_{hoje.strftime('%d%m%Y')}_{datetime.now().strftime('%H%M')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── SEED DEMO (reset para apresentação) ──────────────────────────────────────

@app.post("/api/seed", summary="Reseta banco e recarrega dados de demonstração")
async def seed_demo():
    seed_script = BASE_DIR / "scripts" / "seed_demo.py"
    if not seed_script.exists():
        raise HTTPException(404, "Script de seed não encontrado")

    if IS_VERCEL:
        # No Vercel chamamos a função async diretamente (subprocess não é confiável em serverless)
        try:
            import importlib.util
            spec = importlib.util.spec_from_file_location("seed_demo", seed_script)
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)  # type: ignore[union-attr]
            await mod.seed()
            return {"status": "ok", "output": "Seed executado com sucesso"}
        except Exception as e:
            logger.error(f"[API] Erro no seed (Vercel): {e}")
            raise HTTPException(500, str(e))
    else:
        import subprocess, sys
        try:
            result = subprocess.run(
                [sys.executable, str(seed_script)],
                capture_output=True, text=True, timeout=60
            )
            return {
                "status": "ok" if result.returncode == 0 else "erro",
                "output": result.stdout[-2000:] if result.stdout else "",
                "error":  result.stderr[-1000:] if result.stderr else "",
            }
        except subprocess.TimeoutExpired:
            raise HTTPException(504, "Seed demorou mais de 60s")
        except Exception as e:
            raise HTTPException(500, str(e))


# ── REPROCESSAMENTO DE HISTÓRICO (remessas que nunca passaram pelo pipeline) ──

@app.post(
    "/api/admin/reprocessar-historico",
    summary="Roda Classificador+Montador para remessas que nunca foram planejadas",
)
async def reprocessar_historico(db: AsyncSession = Depends(get_db)):
    """
    Remessas inseridas fora do fluxo real de upload (ex.: seed de teste de
    stress antes desta correção) podem ter ficado com status 'novo' ou
    'planejado' sem nunca terem sido classificadas nem agrupadas em onda.

    Localiza esses grupos por (cd_id, data_extracao) — cada combinação é um
    "dia" de trabalho — e roda o Classificador e o Montador reais para cada
    um, usando data_extracao como data_plano.
    """
    tem_onda = exists().where(OndaRemessa.remessa_id == Remessa.id)

    res_grupos = await db.execute(
        select(Remessa.cd_id, Remessa.data_extracao)
        .where(Remessa.status.in_(["novo", "planejado"]), ~tem_onda)
        .distinct()
    )
    grupos = res_grupos.all()

    res_cds = await db.execute(select(CentroDistribuicao))
    cd_codigo_por_id = {cd.id: cd.codigo for cd in res_cds.scalars().all()}

    classificador = AgenteClassificador(db)
    montador      = AgenteMontador(db)

    dias_processados    = 0
    ondas_criadas        = 0
    remessas_planejadas  = 0

    for cd_id, data_extracao in grupos:
        cd_codigo = cd_codigo_por_id.get(cd_id)
        if not cd_codigo:
            logger.warning(f"[Reprocessamento] CD id={cd_id} não encontrado — grupo ignorado")
            continue

        try:
            # 'planejado' órfão (sem onda) é herança de dados inconsistentes
            # do seed antigo — volta pra 'novo' pro Montador reconsiderá-lo
            # (ele só carrega remessas com status == 'novo').
            await db.execute(
                update(Remessa)
                .where(
                    Remessa.cd_id         == cd_id,
                    Remessa.data_extracao == data_extracao,
                    Remessa.status        == "planejado",
                    ~tem_onda,
                )
                .values(status="novo")
            )
            await classificador.classificar_periodo(cd_id, data_extracao)
            rel_montagem = await montador.montar_plano(cd_codigo, data_extracao)
        except Exception as e:
            # Sessão longa contra o Supabase — uma queda de conexão
            # transitória deixa a Session em PendingRollbackError até um
            # rollback explícito; sem isso, todo grupo seguinte falharia em
            # cascata pelo resto do loop.
            logger.error(f"[Reprocessamento] Falhou {cd_codigo}/{data_extracao}: {e}")
            try:
                await db.rollback()
            except Exception:
                pass  # conexão pode já estar morta — a próxima query force uma nova
            continue

        dias_processados    += 1
        ondas_criadas        += rel_montagem.get("ondas", 0)
        remessas_planejadas  += rel_montagem.get("remessas", 0)

    logger.info(
        f"[Reprocessamento] {dias_processados} dia(s), {ondas_criadas} onda(s), "
        f"{remessas_planejadas} remessa(s) planejada(s)"
    )
    return {
        "dias_processados":    dias_processados,
        "ondas_criadas":       ondas_criadas,
        "remessas_planejadas": remessas_planejadas,
    }


# ── RESET DEMO (limpa dados operacionais, mantém infraestrutura) ─────────────

@app.post("/api/demo/reset", summary="Limpa dados operacionais para demo limpa")
async def reset_demo(db: AsyncSession = Depends(get_db)):
    from sqlalchemy import text
    # Ordem de dependência de FK: historico_eventos.remessa_id referencia
    # remessas, então precisa ser apagada antes (senão viola a FK quando
    # há eventos vinculados a remessas — ex.: eventos de mudança de status).
    tabelas = [
        "onda_remessas", "programacoes_coleta", "ondas",
        "planos_dia", "alertas", "oportunidades_consolidacao",
        "eventos_rastreio", "historico_eventos", "remessas", "uploads",
    ]
    totais = {}
    for t in tabelas:
        res = await db.execute(text(f"DELETE FROM {t}"))
        totais[t] = res.rowcount
    await db.commit()
    logger.info(f"[API] Demo reset: {totais}")
    return {"status": "ok", "deletados": totais}


# ── HEALTH ────────────────────────────────────────────────────────────────────

@app.get("/health")
async def health():
    return {"status": "ok", "version": "1.0.0"}


# ── WEBSOCKET ─────────────────────────────────────────────────────────────────

connected_clients: list[WebSocket] = []

@app.websocket("/ws/dashboard")
async def ws_dashboard(websocket: WebSocket):
    await websocket.accept()
    connected_clients.append(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        connected_clients.remove(websocket)


async def broadcast_dashboard(data: dict):
    import json
    for ws in connected_clients:
        try:
            await ws.send_text(json.dumps(data))
        except Exception:
            pass


# ── FRONTEND (deve ser montado por último) ────────────────────────────────────

if FRONTEND_DIR.exists():
    app.mount("/", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="frontend")
