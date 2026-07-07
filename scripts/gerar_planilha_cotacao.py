# scripts/gerar_planilha_cotacao.py
# Gera o formulário padronizado de cotação de frete — PEO-BD / Becton Dickinson.
#
# Estrutura: Macro Região → Estado → Capital ou Interior
# Exemplo de linha: "Sudeste — São Paulo — SP — Interior — Fracionado"
#
# Um único arquivo enviado para QUALQUER transportadora parceira.
# O sistema importa via POST /api/cotacao/importar.
#
# Uso: python scripts/gerar_planilha_cotacao.py
# Saída: data/outputs/Formulario_Cotacao_BD_<data>.xlsx

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Paleta ────────────────────────────────────────────────────────────────────
AZUL_BD     = "1A3F6F"
BRANCO      = "FFFFFF"
CINZA_CLARO = "F5F5F5"
CINZA_HEAD  = "D6DCE4"
AMARELO     = "FFF2CC"
AZUL_CLARO  = "DBEEF4"
VERDE_CLARO = "E2EFDA"

# Cores por macro região (para highlight visual)
COR_SUDESTE     = "E8F0FE"
COR_SUL         = "E6F4EA"
COR_CENTRO      = "FFF8E1"
COR_NORDESTE    = "FCE4D6"
COR_NORTE       = "F3E5F5"

COR_CAPITAL     = "DBEEF4"   # linha Capital
COR_INTERIOR    = "F5F5F5"   # linha Interior

# ── Definição completa das rotas ──────────────────────────────────────────────
# (macro_regiao, estado, uf, classificacao)
# Uma linha por combinação — FRACIONADO e FTL ficam em colunas separadas
# para facilitar o preenchimento pela transportadora.
#
# Nota: DF não tem "Interior" (é cidade-estado).

ROTAS = [
    # ── SUDESTE ───────────────────────────────────────────────────────────────
    ("Sudeste", "São Paulo",        "SP", "Capital"),
    ("Sudeste", "São Paulo",        "SP", "Interior"),
    ("Sudeste", "Rio de Janeiro",   "RJ", "Capital"),
    ("Sudeste", "Rio de Janeiro",   "RJ", "Interior"),
    ("Sudeste", "Minas Gerais",     "MG", "Capital"),
    ("Sudeste", "Minas Gerais",     "MG", "Interior"),
    ("Sudeste", "Espírito Santo",   "ES", "Capital"),
    ("Sudeste", "Espírito Santo",   "ES", "Interior"),
    # ── SUL ───────────────────────────────────────────────────────────────────
    ("Sul",     "Paraná",           "PR", "Capital"),
    ("Sul",     "Paraná",           "PR", "Interior"),
    ("Sul",     "Santa Catarina",   "SC", "Capital"),
    ("Sul",     "Santa Catarina",   "SC", "Interior"),
    ("Sul",     "Rio Grande do Sul","RS", "Capital"),
    ("Sul",     "Rio Grande do Sul","RS", "Interior"),
    # ── CENTRO-OESTE ──────────────────────────────────────────────────────────
    ("Centro-Oeste", "Distrito Federal",    "DF", "Capital"),
    ("Centro-Oeste", "Goiás",               "GO", "Capital"),
    ("Centro-Oeste", "Goiás",               "GO", "Interior"),
    ("Centro-Oeste", "Mato Grosso",         "MT", "Capital"),
    ("Centro-Oeste", "Mato Grosso",         "MT", "Interior"),
    ("Centro-Oeste", "Mato Grosso do Sul",  "MS", "Capital"),
    ("Centro-Oeste", "Mato Grosso do Sul",  "MS", "Interior"),
    # ── NORDESTE ──────────────────────────────────────────────────────────────
    ("Nordeste", "Bahia",               "BA", "Capital"),
    ("Nordeste", "Bahia",               "BA", "Interior"),
    ("Nordeste", "Pernambuco",          "PE", "Capital"),
    ("Nordeste", "Pernambuco",          "PE", "Interior"),
    ("Nordeste", "Ceará",               "CE", "Capital"),
    ("Nordeste", "Ceará",               "CE", "Interior"),
    ("Nordeste", "Maranhão",            "MA", "Capital"),
    ("Nordeste", "Maranhão",            "MA", "Interior"),
    ("Nordeste", "Rio Grande do Norte", "RN", "Capital"),
    ("Nordeste", "Rio Grande do Norte", "RN", "Interior"),
    ("Nordeste", "Paraíba",             "PB", "Capital"),
    ("Nordeste", "Paraíba",             "PB", "Interior"),
    ("Nordeste", "Alagoas",             "AL", "Capital"),
    ("Nordeste", "Alagoas",             "AL", "Interior"),
    ("Nordeste", "Sergipe",             "SE", "Capital"),
    ("Nordeste", "Sergipe",             "SE", "Interior"),
    ("Nordeste", "Piauí",               "PI", "Capital"),
    ("Nordeste", "Piauí",               "PI", "Interior"),
    # ── NORTE ─────────────────────────────────────────────────────────────────
    ("Norte", "Pará",       "PA", "Capital"),
    ("Norte", "Pará",       "PA", "Interior"),
    ("Norte", "Amazonas",   "AM", "Capital"),
    ("Norte", "Amazonas",   "AM", "Interior"),
    ("Norte", "Rondônia",   "RO", "Capital"),
    ("Norte", "Rondônia",   "RO", "Interior"),
    ("Norte", "Roraima",    "RR", "Capital"),
    ("Norte", "Roraima",    "RR", "Interior"),
    ("Norte", "Acre",       "AC", "Capital"),
    ("Norte", "Acre",       "AC", "Interior"),
    ("Norte", "Amapá",      "AP", "Capital"),
    ("Norte", "Amapá",      "AP", "Interior"),
    ("Norte", "Tocantins",  "TO", "Capital"),
    ("Norte", "Tocantins",  "TO", "Interior"),
]

# ── Colunas (1-based; B=2) ────────────────────────────────────────────────────
# Pré-preenchidas pela BD (não editáveis pela transportadora)
COL_MACRO      = 2   # B — Macro Região
COL_ESTADO     = 3   # C — Estado
COL_UF         = 4   # D — UF (sigla)
COL_CLASS      = 5   # E — Capital / Interior

# Campos a preencher pela transportadora (células amarelas)
COL_COBRE      = 6   # F — Sim / Não
COL_PRECO_KG   = 7   # G — Fracionado: R$/kg
COL_PESO_MIN   = 8   # H — Fracionado: peso mínimo kg
COL_PRECO_MIN  = 9   # I — Fracionado: preço mínimo R$
COL_PRAZO_FRAC = 10  # J — Fracionado: prazo (d.u.)
COL_PRECO_FTL  = 11  # K — FTL: R$/viagem
COL_PRAZO_FTL  = 12  # L — FTL: prazo (d.u.)
COL_ADVALOREM  = 13  # M — Ad Valorem % NF
COL_GRIS       = 14  # N — GRIS % NF
COL_SLA        = 15  # O — SLA confirmação (h)
COL_OBS        = 16  # P — Observações

LARGURAS = [2, 16, 22, 6, 12, 10, 13, 13, 15, 13, 15, 13, 13, 10, 12, 32]

# Linhas de layout
HEADER_ROW = 14   # Linha do cabeçalho da tabela
DATA_START  = 15  # Primeira linha de dados

# Células fixas lidas pelo importador
CEL_NOME_TRANSP  = (3, 3)   # C3
CEL_CNPJ         = (3, 8)   # H3
CEL_VALIDADE_INI = (4, 3)   # C4
CEL_VALIDADE_FIM = (4, 8)   # H4

# Capas de macroregião para cor de fundo
COR_MACRO = {
    "Sudeste":      COR_SUDESTE,
    "Sul":          COR_SUL,
    "Centro-Oeste": COR_CENTRO,
    "Nordeste":     COR_NORDESTE,
    "Norte":        COR_NORTE,
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def _fill(c): return PatternFill("solid", fgColor=c)
def _borda(cor="BBBBBB"):
    s = Side(border_style="thin", color=cor)
    return Border(left=s, right=s, top=s, bottom=s)
def _borda_m(cor="888888"):
    s = Side(border_style="medium", color=cor)
    return Border(left=s, right=s, top=s, bottom=s)
def _aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def gerar_formulario(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotacao_Frete"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"B{DATA_START}"   # congela cabeçalho

    for i, larg in enumerate(LARGURAS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg

    # ── Linha 1: Título ──────────────────────────────────────────────────────
    ws.merge_cells("B1:P1")
    ws["B1"].value     = "FORMULÁRIO DE COTAÇÃO DE FRETE — BECTON DICKINSON"
    ws["B1"].font      = Font(bold=True, size=15, color=BRANCO)
    ws["B1"].fill      = _fill(AZUL_BD)
    ws["B1"].alignment = _aln("center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("B2:P2")
    ws["B2"].value = (
        f"Preencha as células em AMARELO e devolva à BD  |  "
        f"Emitido em {datetime.now().strftime('%d/%m/%Y')}  |  "
        "Duvidas: logistica@bd.com  |  (11) 3000-0000"
    )
    ws["B2"].font      = Font(italic=True, size=9, color="444444")
    ws["B2"].fill      = _fill(AZUL_CLARO)
    ws["B2"].alignment = _aln("center")
    ws.row_dimensions[2].height = 16

    # ── Linhas 3-4: Identificação da transportadora ──────────────────────────
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20

    for lbl1, txt1, mrg1, lbl2, txt2, mrg2 in [
        ("B3", "TRANSPORTADORA:",    "C3:G3", "H3", "CNPJ:",         "I3:P3"),
        ("B4", "RESPONSAVEL/CARGO:", "C4:G4", "H4", "VALIDADE ATE:", "I4:P4"),
    ]:
        ws[lbl1].value     = txt1
        ws[lbl1].font      = Font(bold=True, size=10, color=AZUL_BD)
        ws[lbl1].alignment = _aln("right")
        ws.merge_cells(mrg1)
        ws[mrg1.split(":")[0]].fill   = _fill(AMARELO)
        ws[mrg1.split(":")[0]].border = _borda()

        ws[lbl2].value     = txt2
        ws[lbl2].font      = Font(bold=True, size=10, color=AZUL_BD)
        ws[lbl2].alignment = _aln("right")
        ws.merge_cells(mrg2)
        ws[mrg2.split(":")[0]].fill   = _fill(AMARELO)
        ws[mrg2.split(":")[0]].border = _borda()

    # ── Linhas 5-7: Espaço + legenda ─────────────────────────────────────────
    ws.row_dimensions[5].height = 8

    leg = [
        ("B6:E6", "  Regiao/Estado/UF = pré-preenchido pela BD — nao altere",  CINZA_HEAD),
        ("F6:F6", "  Cobre? = Sim ou Nao (obrigatorio)",                         AMARELO),
        ("G6:J6", "  Fracionado: preencha G, H, I, J",                          AMARELO),
        ("K6:L6", "  FTL: preencha K, L",                                       AMARELO),
        ("M6:P6", "  Ad Valorem / GRIS / SLA / Obs",                            AMARELO),
    ]
    ws.row_dimensions[6].height = 14
    for mrg, txt, cor in leg:
        ws.merge_cells(mrg)
        c = ws[mrg.split(":")[0]]
        c.value     = txt
        c.font      = Font(italic=True, size=8, color="333333")
        c.fill      = _fill(cor)
        c.alignment = _aln("left")

    # ── Linhas 7-12: Instruções ───────────────────────────────────────────────
    ws.row_dimensions[7].height = 8

    instrucoes = [
        ("INSTRUCOES:",
         "Preencha SOMENTE as celulas em AMARELO. "
         "Nao altere as colunas B, C, D, E (pre-preenchidas pela BD)."),
        ("Coluna F (Cobre?):",
         "Selecione SIM se atende a rota; NAO se nao atende. "
         "Se NAO, deixe os demais campos em branco."),
        ("Colunas G-J (Fracionado):",
         "Preco/kg, peso minimo, preco minimo por coleta e prazo. "
         "Peso cubado = volume (m3) x 300. Cobrar o maior entre real e cubado."),
        ("Colunas K-L (FTL):",
         "Preco por viagem de carga fechada (caminhao 30 m3/10.000 kg) e prazo. "
         "Deixe em branco se nao trabalha com FTL nesta rota."),
        ("Colunas M-N (Ad Valorem / GRIS):",
         "Informar como percentual decimal: 0,1500 = 0,15% sobre o valor total da NF."),
        ("Devolucao:",
         "Envie o arquivo preenchido e assinado para logistica@bd.com com assunto "
         "'COTACAO FRETE BD [nome da empresa] [mes/ano]'."),
    ]

    for idx, (chave, texto) in enumerate(instrucoes, start=8):
        ws.row_dimensions[idx].height = 13
        ws[f"B{idx}"].value     = chave
        ws[f"B{idx}"].font      = Font(bold=(chave == "INSTRUCOES:"), size=8,
                                       color=AZUL_BD if chave == "INSTRUCOES:" else "444444")
        ws[f"B{idx}"].alignment = _aln("left")
        ws.merge_cells(f"C{idx}:P{idx}")
        ws[f"C{idx}"].value     = texto
        ws[f"C{idx}"].font      = Font(size=8, color="333333")
        ws[f"C{idx}"].alignment = _aln("left")

    ws.row_dimensions[13].height = 6

    # ── Linha 14: Cabeçalho da tabela ─────────────────────────────────────────
    headers = [
        "",
        "Macro\nRegiao",
        "Estado",
        "UF",
        "Capital /\nInterior",
        "Cobre?\n(Sim/Nao)",
        # Fracionado
        "Frac.\nPreco/kg (R$)",
        "Frac.\nPeso min (kg)",
        "Frac.\nPreco min (R$)",
        "Frac.\nPrazo (d.u.)",
        # FTL
        "FTL\nPreco/viagem (R$)",
        "FTL\nPrazo (d.u.)",
        # Encargos
        "Ad Valorem\n(% s/ NF)",
        "GRIS\n(% s/ NF)",
        "SLA confirm.\ncoleta (h)",
        "Observacoes",
    ]
    ws.row_dimensions[HEADER_ROW].height = 40
    for col_num, cab in enumerate(headers, start=1):
        if col_num == 1:
            continue
        c = ws.cell(row=HEADER_ROW, column=col_num, value=cab)
        c.fill      = _fill(AZUL_BD)
        c.font      = Font(bold=True, size=9, color=BRANCO)
        c.alignment = _aln("center", wrap=True)
        c.border    = _borda("999999")

    # Sub-cabeçalho de grupo (Fracionado / FTL)
    # Não adicionamos linha extra — o wrap no header já deixa claro

    # ── Linhas de dados ───────────────────────────────────────────────────────
    n_rows = len(ROTAS)
    dv_range = f"F{DATA_START}:F{DATA_START + n_rows - 1}"
    dv = DataValidation(type="list", formula1='"Sim,Nao"',
                        allow_blank=False, showDropDown=False)
    dv.sqref = dv_range
    ws.add_data_validation(dv)

    macro_ant  = None
    estado_ant = None

    for idx, (macro, estado, uf, classif) in enumerate(ROTAS):
        row         = DATA_START + idx
        ws.row_dimensions[row].height = 18

        eh_nova_macro  = macro  != macro_ant
        eh_novo_estado = estado != estado_ant
        macro_ant  = macro
        estado_ant = estado

        cor_macro  = COR_MACRO.get(macro, CINZA_CLARO)
        cor_linha  = COR_CAPITAL if classif == "Capital" else COR_INTERIOR

        for col in range(2, 17):
            c = ws.cell(row=row, column=col)
            c.border    = _borda()
            c.alignment = _aln("center")

            if col == COL_MACRO:
                c.value     = macro if eh_nova_macro else ""
                c.font      = Font(bold=True, size=9, color=AZUL_BD)
                c.fill      = _fill(cor_macro)
                c.alignment = _aln("left")

            elif col == COL_ESTADO:
                c.value     = estado if eh_novo_estado else ""
                c.font      = Font(size=9, color="222222")
                c.fill      = _fill(cor_macro)
                c.alignment = _aln("left")

            elif col == COL_UF:
                c.value = uf
                c.font  = Font(bold=True, size=9)
                c.fill  = _fill(cor_macro)

            elif col == COL_CLASS:
                c.value = classif
                c.font  = Font(size=9,
                               bold=(classif == "Capital"),
                               color="1A3F6F" if classif == "Capital" else "444444")
                c.fill  = _fill(cor_linha)

            elif col == COL_COBRE:
                c.fill = _fill(AMARELO)
                c.font = Font(size=10)

            elif col in (COL_PRECO_KG, COL_PRECO_MIN, COL_PRECO_FTL):
                c.fill          = _fill(AMARELO)
                c.number_format = '#,##0.00'

            elif col == COL_PESO_MIN:
                c.fill          = _fill(AMARELO)
                c.number_format = '#,##0'

            elif col in (COL_PRAZO_FRAC, COL_PRAZO_FTL, COL_SLA):
                c.fill          = _fill(AMARELO)
                c.number_format = '0'

            elif col in (COL_ADVALOREM, COL_GRIS):
                c.fill          = _fill(AMARELO)
                c.number_format = '0.0000'

            elif col == COL_OBS:
                c.fill      = _fill(AMARELO)
                c.alignment = _aln("left")

        # Borda mais forte na transição entre macro regiões
        if eh_nova_macro and idx > 0:
            borda_top = Side(border_style="medium", color="555555")
            for col in range(2, 17):
                c = ws.cell(row=row, column=col)
                existing = c.border
                c.border = Border(
                    left=existing.left, right=existing.right,
                    bottom=existing.bottom, top=borda_top
                )

    # ── Rodapé ───────────────────────────────────────────────────────────────
    row_foot = DATA_START + n_rows + 1
    ws.merge_cells(f"B{row_foot}:P{row_foot}")
    ws[f"B{row_foot}"].value = (
        "Peso cubado = Volume (m3) x 300 | Cobrar o MAIOR entre peso real e peso cubado | "
        "FTL = caminhao 30 m3 / 10.000 kg dedicado | "
        "Ad Valorem e GRIS em decimal: 0,1500 = 0,15%"
    )
    ws[f"B{row_foot}"].font      = Font(italic=True, size=8, color="555555")
    ws[f"B{row_foot}"].alignment = _aln("left")
    ws.row_dimensions[row_foot].height = 13

    row_ass = row_foot + 2
    ws.row_dimensions[row_ass].height = 26
    for lbl_c, val_c, txt in [
        ("B", "C", "Assinatura:"),
        ("H", "I", "Carimbo / Data:"),
        ("M", "N", "Validade confirmada:"),
    ]:
        ws[f"{lbl_c}{row_ass}"].value     = txt
        ws[f"{lbl_c}{row_ass}"].font      = Font(size=10)
        ws[f"{lbl_c}{row_ass}"].alignment = _aln("right")
        ws[f"{val_c}{row_ass}"].border    = Border(
            bottom=Side(border_style="medium", color="000000")
        )
        ws.merge_cells(f"{val_c}{row_ass}:{chr(ord(val_c)+3)}{row_ass}")

    # ── Aba de referência: capitais por UF ───────────────────────────────────
    ws_ref = wb.create_sheet("Referencia_Capitais")
    ws_ref.sheet_view.showGridLines = False
    ws_ref.column_dimensions["B"].width = 6
    ws_ref.column_dimensions["C"].width = 30
    ws_ref.column_dimensions["D"].width = 30

    ws_ref.merge_cells("B1:D1")
    ws_ref["B1"].value     = "REFERENCIA: Capital de cada Estado (usado pelo sistema)"
    ws_ref["B1"].font      = Font(bold=True, size=11, color=BRANCO)
    ws_ref["B1"].fill      = _fill(AZUL_BD)
    ws_ref["B1"].alignment = _aln("center")
    ws_ref.row_dimensions[1].height = 22

    capitais = [
        ("UF", "Estado", "Capital"),
        ("AC", "Acre",                "Rio Branco"),
        ("AL", "Alagoas",             "Maceió"),
        ("AP", "Amapá",               "Macapá"),
        ("AM", "Amazonas",            "Manaus"),
        ("BA", "Bahia",               "Salvador"),
        ("CE", "Ceará",               "Fortaleza"),
        ("DF", "Distrito Federal",    "Brasília"),
        ("ES", "Espírito Santo",      "Vitória"),
        ("GO", "Goiás",               "Goiânia"),
        ("MA", "Maranhão",            "São Luís"),
        ("MT", "Mato Grosso",         "Cuiabá"),
        ("MS", "Mato Grosso do Sul",  "Campo Grande"),
        ("MG", "Minas Gerais",        "Belo Horizonte"),
        ("PA", "Pará",                "Belém"),
        ("PB", "Paraíba",             "João Pessoa"),
        ("PR", "Paraná",              "Curitiba"),
        ("PE", "Pernambuco",          "Recife"),
        ("PI", "Piauí",               "Teresina"),
        ("RJ", "Rio de Janeiro",      "Rio de Janeiro"),
        ("RN", "Rio Grande do Norte", "Natal"),
        ("RS", "Rio Grande do Sul",   "Porto Alegre"),
        ("RO", "Rondônia",            "Porto Velho"),
        ("RR", "Roraima",             "Boa Vista"),
        ("SC", "Santa Catarina",      "Florianópolis"),
        ("SP", "São Paulo",           "São Paulo"),
        ("SE", "Sergipe",             "Aracaju"),
        ("TO", "Tocantins",           "Palmas"),
    ]
    for i, (uf, estado, capital) in enumerate(capitais, start=2):
        ws_ref.row_dimensions[i].height = 16
        for col, val in [(2, uf), (3, estado), (4, capital)]:
            c = ws_ref.cell(row=i, column=col, value=val)
            c.border    = _borda()
            c.font      = Font(bold=(i == 2), size=10)
            c.alignment = _aln("center" if col == 2 else "left")
            if i == 2:
                c.fill = _fill(CINZA_HEAD)

    # ── Aba com instruções detalhadas ────────────────────────────────────────
    ws_inst = wb.create_sheet("Instrucoes_Detalhadas")
    ws_inst.sheet_view.showGridLines = False
    ws_inst.column_dimensions["A"].width = 3
    ws_inst.column_dimensions["B"].width = 30
    ws_inst.column_dimensions["C"].width = 68

    ws_inst.merge_cells("B1:C1")
    ws_inst["B1"].value     = "INSTRUCOES DETALHADAS — FORMULARIO DE COTACAO BD"
    ws_inst["B1"].font      = Font(bold=True, size=13, color=BRANCO)
    ws_inst["B1"].fill      = _fill(AZUL_BD)
    ws_inst["B1"].alignment = _aln("center")
    ws_inst.row_dimensions[1].height = 28

    blocos_inst = [
        ("ESTRUTURA DA TABELA", [
            ("Macro Região",      "Agrupamento geográfico: Sudeste, Sul, Centro-Oeste, Nordeste, Norte."),
            ("Estado",            "Estado da Federação (26 estados + Distrito Federal)."),
            ("UF",                "Sigla do estado (ex.: SP, RJ, MG). Usada pelo sistema para vincular ao cliente."),
            ("Capital / Interior","'Capital' = cidade capital do estado e região metropolitana. "
                                  "'Interior' = demais municípios do estado. "
                                  "O sistema BD determina a classificação pelo endereço do cliente."),
        ]),
        ("CAMPOS A PREENCHER", [
            ("F — Cobre? (Sim/Nao)",     "Obrigatorio. Se nao atende a rota, coloque NAO e deixe os demais campos em branco."),
            ("G — Preco/kg Frac. (R$)",  "Tarifa por kg para frete fracionado. Aplicar ao maior peso entre real e cubado."),
            ("H — Peso min. (kg)",       "Peso minimo faturado por coleta fracionado."),
            ("I — Preco min. Frac. (R$)","Valor minimo cobrado por coleta fracionado, independente do peso."),
            ("J — Prazo Frac. (d.u.)",   "Dias uteis para entrega no destinatario, a partir da coleta no CD."),
            ("K — Preco FTL/viagem (R$)","Preco fixo por viagem de carga fechada (caminhao dedicado 30 m3/10 t)."),
            ("L — Prazo FTL (d.u.)",     "Dias uteis para entrega em carga fechada."),
            ("M — Ad Valorem (% NF)",    "Cobertura de avaria e roubo. Informar como decimal: 0,1500 = 0,15%."),
            ("N — GRIS (% NF)",          "Gerenciamento de Risco. Informar como decimal: 0,1000 = 0,10%."),
            ("O — SLA confirm. (h)",     "Horas para confirmar coleta apos receber programacao da BD. Zero = imediato."),
            ("P — Observacoes",          "Condicoes especiais: pedagio incluso, restricao de horario, sub-regioes, etc."),
        ]),
        ("DEVOLUCAO", [
            ("Prazo",    "Devolver em ate 5 dias uteis da data de emissao."),
            ("E-mail",   "logistica@bd.com"),
            ("Assunto",  "COTACAO FRETE BD [nome da empresa] [mes/ano]"),
            ("Formato",  "Devolver o arquivo .xlsx preenchido, assinado e carimbado. Nao renomear o arquivo."),
        ]),
    ]

    linha = 3
    for titulo, itens in blocos_inst:
        ws_inst.row_dimensions[linha].height = 8
        linha += 1
        ws_inst.merge_cells(f"B{linha}:C{linha}")
        ws_inst[f"B{linha}"].value     = titulo
        ws_inst[f"B{linha}"].font      = Font(bold=True, size=11, color=BRANCO)
        ws_inst[f"B{linha}"].fill      = _fill(AZUL_BD)
        ws_inst[f"B{linha}"].alignment = _aln("left")
        ws_inst.row_dimensions[linha].height = 20
        linha += 1

        for campo, desc in itens:
            ws_inst.row_dimensions[linha].height = 15
            ws_inst[f"B{linha}"].value     = f"  {campo}"
            ws_inst[f"B{linha}"].font      = Font(bold=True, size=10)
            ws_inst[f"B{linha}"].alignment = _aln("left", wrap=True)
            ws_inst[f"C{linha}"].value     = desc
            ws_inst[f"C{linha}"].font      = Font(size=10)
            ws_inst[f"C{linha}"].alignment = _aln("left", wrap=True)
            ws_inst.row_dimensions[linha].height = max(15, len(desc) // 5)
            linha += 1

    wb.active = wb["Cotacao_Frete"]
    nome    = f"Formulario_Cotacao_BD_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    caminho = output_dir / nome
    wb.save(caminho)
    return caminho


def main():
    output_dir = Path(__file__).parent.parent / "data" / "outputs"
    output_dir.mkdir(parents=True, exist_ok=True)
    caminho = gerar_formulario(output_dir)
    print(f"Formulario gerado: {caminho}")
    print(f"Total de rotas: {len(ROTAS)} linhas")


if __name__ == "__main__":
    main()
